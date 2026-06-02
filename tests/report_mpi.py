"""
report_mpi.py  --  Traffic Automaton MPI benchmark report generator

Estructura del reporte
----------------------
  Hoja 0 — Secuenciales
      Compara todas las variantes secuenciales presentes en el CSV y
      determina la más rápida, que actúa como referencia base para el
      resto del análisis.

  Hoja 1 — MPI sin optimización de compilador
      Referencia secuencial óptima vs MPI sin flags de optimización,
      para cada conteo de procesos disponible en los datos.

  Hoja 2 — MPI con optimización de compilador
      Referencia secuencial óptima vs MPI con flags de optimización
      (-O3 -march=native -flto …), para cada conteo de procesos.

  Hoja 3 — Comparación final
      Referencia secuencial óptima vs el mejor resultado de MPI sin
      optimización vs el mejor resultado de MPI con optimización.

  Hoja 4 — Experimento de densidad
      Tiempo de ejecución y velocidad media del autómata en función de
      la densidad de tráfico, para cada configuración de procesos.

Cada hoja incluye:
  - Tabla de mediciones individuales por repetición (ms)
  - Tabla de promedios con speedup relativo a la referencia secuencial
  - Gráfica de tiempo promedio vs tamaño de carretera (escala log)
  - Gráfica de speedup vs tamaño de carretera
  - Gráfica de escalabilidad fuerte (speedup y eficiencia vs número
    de procesos, por nivel de memoria)

CSV esperado (salida de bench_traffic_mpi.sh):
  impl, np, road_length, density, repetition, wall_time_ms, avg_velocity

Convención de la columna np:
  0   → variante secuencial (sin MPI)
  ≥1  → MPI con ese número de procesos

Uso:
  python report_mpi.py [directorio_resultados]
  python report_mpi.py tests/mpi/results_traffic_mpi
"""

from __future__ import annotations
import math
import os
import sys
from dataclasses import dataclass

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# Configuración
# =============================================================================

BASE_DIR = (sys.argv[1].rstrip("/") if len(sys.argv) > 1
               else "tests/mpi")
TRAFFIC = "results_traffic"
TRAFFIC_MPI = "results_traffic_mpi"
RESULTS_DIR = os.path.join(BASE_DIR, TRAFFIC)
RESULTS_DIR_MPI = os.path.join(BASE_DIR, TRAFFIC_MPI)
OUTPUT_PATH = os.path.join(RESULTS_DIR_MPI, "reporte_traffic_mpi.xlsx")
CHARTS_DIR  = os.path.join(RESULTS_DIR_MPI, "charts")
CSV_FILE_SEQ   = os.path.join(RESULTS_DIR, "data.csv")
CSV_FILE_MPI    = os.path.join(RESULTS_DIR_MPI, "data.csv")

SCALING_DENSITY      = 0.50
N_VALUES             = [32000, 256000, 2000000, 8000000, 20000000, 40000000]
DENSITY_EXPERIMENT_N = 80000000
DENSITY_VALUES       = [0.10, 0.30, 0.50, 0.70, 0.90]

# =============================================================================
# Nombres en prosa (sin abreviaciones)
# =============================================================================

NOMBRE_SEQ         = "Serial Base"
NOMBRE_SEQ_OPT     = "Serial con Optimización de Compilador"
NOMBRE_SEQ_MEM     = "Serial con Optimización de Memoria"
NOMBRE_SEQ_MEM_OPT = "Serial con Optimización de Compilador y Memoria"
NOMBRE_MPI         = "MPI sin Optimización de Compilador"
NOMBRE_MPI_OPT     = "MPI con Optimización de Compilador"

IMPL_NAME_MAP = {
    "traffic_seq":         NOMBRE_SEQ,
    "traffic_seq_opt":     NOMBRE_SEQ_OPT,
    "traffic_seq_mem":     NOMBRE_SEQ_MEM,
    "traffic_seq_mem_opt": NOMBRE_SEQ_MEM_OPT,
    "traffic_mpi":         NOMBRE_MPI,
    "traffic_mpi_opt":     NOMBRE_MPI_OPT,
}

ALL_SERIAL_NAMES = {NOMBRE_SEQ, NOMBRE_SEQ_OPT, NOMBRE_SEQ_MEM, NOMBRE_SEQ_MEM_OPT}

# =============================================================================
# Paletas de color
# =============================================================================

PALETTE = {
    NOMBRE_SEQ:         "#2E75B6",
    NOMBRE_SEQ_OPT:     "#ED7D31",
    NOMBRE_SEQ_MEM:     "#70AD47",
    NOMBRE_SEQ_MEM_OPT: "#C00000",
    NOMBRE_MPI:         "#7030A0",
    NOMBRE_MPI_OPT:     "#00B0F0",
}

NP_PALETTE = {
   NOMBRE_MPI: {
      1: "#2E75B6",
      2: "#ED7D31",
      3: "#30D84C",
      4: "#70AD47",
      5: "#C00000",
      6: "#7030A0",
   },
   NOMBRE_MPI_OPT: {
      1: "#30A1CE",
      2: "#ED3131",
      3: "#288569",
      4: "#6F3EBE",
      5: "#C1F336",
      6: "#C226A8",
   },
}

CHART_STYLE = {
    "figure.facecolor":  "white",
    "axes.facecolor":    "white",
    "axes.grid":         True,
    "grid.color":        "#E7EDF3",
    "grid.linestyle":    "--",
    "grid.alpha":        0.75,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.titleweight":  "bold",
    "axes.labelsize":    10,
    "axes.titlesize":    11,
    "legend.frameon":    False,
    "font.size":         10,
}

FONT_NAME = "Calibri"

C = {
    "dark":       "1F3557",
    "mid":        "2E75B6",
    "light":      "DCE6F1",
    "alt":        "F4F8FC",
    "green_bg":   "E2F0D9",
    "green_fg":   "2F6B1E",
    "summary_bg": "EBF3FB",
    "summary_fg": "1F4E79",
    "impl_hdr":   "1F4E79",
    "sep":        "D9E1F2",
}

# =============================================================================
# Tipos de datos
# =============================================================================

@dataclass(frozen=True)
class Impl:
    """Identifica una implementación por nombre y número de procesos MPI."""
    name: str
    np: int  # 0 = secuencial sin MPI

    @property
    def label(self) -> str:
        if self.np == 0:
            return self.name
        proc = "proceso" if self.np == 1 else "procesos"
        return f"{self.name}  ({self.np} {proc})"

    @property
    def short_label(self) -> str:
        if self.np == 0:
            return self.name
        return f"{self.name}_np{self.np}"

    @property
    def color(self) -> str:
        if self.np == 0:
            return PALETTE.get(self.name, "#888888")
        color = NP_PALETTE.get(self.name, {}).get(self.np, "#888888")
        return color

# =============================================================================
# Funciones auxiliares
# =============================================================================

def _size_label(n: int) -> str:
    if n >= 1_000_000 and n % 1_000_000 == 0:
        return f"{n // 1_000_000}M"
    if n >= 1_000_000:
        return f"{n / 1_000_000:.1f}M"
    if n >= 1_000 and n % 1_000 == 0:
        return f"{n // 1_000}K"
    return str(n)

def _mean(vals: list) -> float:
    clean = [float(v) for v in vals if v is not None and not math.isnan(float(v))]
    return sum(clean) / len(clean) if clean else math.nan

def _std(vals: list) -> float:
    clean = [float(v) for v in vals if v is not None and not math.isnan(float(v))]
    if len(clean) < 2:
        return 0.0
    m = _mean(clean)
    return math.sqrt(sum((v - m) ** 2 for v in clean) / len(clean))

def _cv(vals: list) -> float:
    m = _mean(vals)
    return 100 * _std(vals) / m if m and not math.isnan(m) else math.nan

# =============================================================================
# Carga y preparación de datos
# =============================================================================

_CSV_CACHE: dict[str, pd.DataFrame] = {}


def load_csv(csv_path: str) -> pd.DataFrame | None:
    if csv_path in _CSV_CACHE:
        return _CSV_CACHE[csv_path]

    if not os.path.exists(csv_path):
        print(f"  [--] {csv_path} no encontrado.")
        return None

    df = pd.read_csv(csv_path)
    df.columns = [c.strip().lower() for c in df.columns]

    # Unificar nombres de columnas entre CSV secuencial y CSV MPI.
    # Algunos benchmarks generan "threads" y otros "np" para el grado de paralelismo.
    rename_map = {}
    if "parallelism" not in df.columns:
        if "np" in df.columns:
            rename_map["np"] = "parallelism"
        elif "threads" in df.columns:
            rename_map["threads"] = "parallelism"
    if rename_map:
        df = df.rename(columns=rename_map)

    required = {"impl", "parallelism", "road_length", "density", "repetition", "wall_time_ms", "avg_velocity"}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"{csv_path} no contiene las columnas requeridas: {missing}")

    # Normalizar tipos
    df["impl"] = df["impl"].astype(str).str.strip().map(lambda x: IMPL_NAME_MAP.get(x, x))
    df["parallelism"] = pd.to_numeric(df["parallelism"], errors="coerce").fillna(0).astype(int)
    df["road_length"] = pd.to_numeric(df["road_length"], errors="coerce").astype("Int64").astype(int)
    df["density"] = pd.to_numeric(df["density"], errors="coerce").astype(float)
    df["repetition"] = pd.to_numeric(df["repetition"], errors="coerce").astype("Int64").astype(int)
    df["wall_time_ms"] = pd.to_numeric(df["wall_time_ms"], errors="coerce").astype(float)
    df["avg_velocity"] = pd.to_numeric(df["avg_velocity"], errors="coerce").astype(float)

    # Marcar mediciones fallidas como NaN
    df.loc[df["wall_time_ms"] == 0, "wall_time_ms"] = math.nan

    _CSV_CACHE[csv_path] = df
    print(f"  [OK] {csv_path}  ({len(df)} filas)")
    return df


def load_benchmark_data(*csv_paths: str) -> pd.DataFrame:
    frames = [df for path in csv_paths if (df := load_csv(path)) is not None]
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)

    # Evitar duplicados exactos si el pipeline escribió el mismo registro dos veces.
    df = df.drop_duplicates(
        subset=["impl", "parallelism", "road_length", "density", "repetition", "wall_time_ms", "avg_velocity"]
    ).reset_index(drop=True)
    return df



def get_scaling_df(df: pd.DataFrame) -> pd.DataFrame:
    return df[
        df["road_length"].isin(N_VALUES) & (df["density"] == SCALING_DENSITY)
    ].copy()


def get_density_df(df: pd.DataFrame) -> pd.DataFrame:
    return df[df["road_length"] == DENSITY_EXPERIMENT_N].copy()


def build_all_reps(df: pd.DataFrame) -> dict:
    """Agrupa las repeticiones por (Impl, road_length)."""
    result = {}
    for (name, par), grp in df.groupby(["impl", "parallelism"]):
        impl = Impl(name=str(name), np=int(par))
        result[impl] = {}
        for size, sub in grp.groupby("road_length"):
            result[impl][int(size)] = sorted(
                [
                    (int(r), float(t), float(v))
                    for r, t, v in zip(
                        sub["repetition"], sub["wall_time_ms"], sub["avg_velocity"]
                    )
                    if not math.isnan(float(t))
                ],
                key=lambda x: x[0],
            )
    return result


def compute_avgs(all_reps: dict) -> dict:
    """Calcula el tiempo promedio por (Impl, road_length)."""
    return {
        impl: {s: _mean([t for _, t, _ in pairs]) for s, pairs in sizes.items()}
        for impl, sizes in all_reps.items()
    }


def find_best_serial(avg_data: dict, sizes: list) -> Impl | None:
    """Devuelve la variante secuencial con menor tiempo acumulado."""
    best_impl, best_sum = None, float("inf")
    for impl, times in avg_data.items():
        if impl.name not in ALL_SERIAL_NAMES or impl.np != 0:
            continue
        total = sum(times.get(s, float("inf")) for s in sizes)
        if total < best_sum:
            best_sum = total
            best_impl = impl
    return best_impl


def get_mpi_variants(avg_data: dict, mpi_name: str) -> list[Impl]:
    """Devuelve todos los (Impl, np≥1) de una familia MPI, ordenados por np."""
    return sorted(
        [i for i in avg_data if i.name == mpi_name and i.np >= 1],
        key=lambda i: i.np,
    )


def best_mpi_variant(family: str, avg_data: dict, best_seq: Impl, sizes: list) -> Impl | None:
    """Devuelve el número de procesos que maximiza el speedup promedio."""
    candidates = [i for i in avg_data if i.name == family and i.np >= 1]
    if not candidates:
        return None
    ref = avg_data.get(best_seq, {})
    best_impl, best_sp = None, 0.0
    for impl in candidates:
        sp_vals = [
            ref[s] / avg_data[impl][s]
            for s in sizes
            if s in avg_data[impl] and s in ref and avg_data[impl][s] > 0
        ]
        sp = _mean(sp_vals)
        if sp > best_sp:
            best_sp = sp
            best_impl = impl
    return best_impl

# =============================================================================
# Estilos de Excel
# =============================================================================

def _thin_border():
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    thick = Side(style="medium", color="1F4E79")
    thin  = Side(style="thin",   color="BDD7EE")
    return Border(left=thick, right=thick, top=thin, bottom=thick)

def _hdr(cell, v, bg=C["mid"], fg="FFFFFF", size=10, bold=True, align="center"):
    cell.value     = v
    cell.font      = Font(name=FONT_NAME, bold=bold, size=size, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = _thin_border()

def _dat(cell, v, fmt=None, bg="FFFFFF", fg="000000", bold=False, align="right"):
    cell.value     = v
    cell.font      = Font(name=FONT_NAME, bold=bold, size=10, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _thin_border()
    if fmt:
        cell.number_format = fmt

def _summary_dat(cell, v, fmt=None, bold=False):
    _dat(cell, v, fmt=fmt, bg=C["summary_bg"], fg=C["summary_fg"], bold=bold)

def _set_w(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def _title_row(ws, text, cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["dark"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26

# =============================================================================
# Funciones de gráficas
# =============================================================================

def _save_fig(fig, name: str) -> str:
    os.makedirs(CHARTS_DIR, exist_ok=True)
    path = os.path.join(CHARTS_DIR, name)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return path

def _categorical_xticks(ax, sizes: list):
    positions = list(range(len(sizes)))
    ax.set_xticks(positions)
    ax.set_xticklabels([_size_label(s) for s in sizes], rotation=20, ha="right")
    return positions


def chart_time(impls: list, avg_data: dict, sizes: list,
               title: str, fname: str) -> str:
    """Tiempo promedio (ms) vs tamaño de carretera N en escala logarítmica."""
    positions = list(range(len(sizes)))
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 5))
        for impl in impls:
            times = avg_data.get(impl, {})
            xs = [positions[i] for i, s in enumerate(sizes) if s in times]
            ys = [times[s] for s in sizes if s in times]
            if xs:
                ax.plot(xs, ys, marker="o", lw=2.2, color=impl.color, label=impl.label)
        _categorical_xticks(ax, sizes)
        ax.set_yscale("log")
        ax.set_ylabel("Tiempo promedio (ms)  [escala log]")
        ax.set_xlabel("Tamaño de la carretera N")
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
    return _save_fig(fig, fname)


def chart_speedup(ref_impl: Impl, cmp_impls: list, avg_data: dict, sizes: list,
                  title: str, fname: str, note: str | None = None) -> str:
    """Speedup relativo a la referencia secuencial vs tamaño de carretera."""
    ref_avgs  = avg_data.get(ref_impl, {})
    positions = list(range(len(sizes)))
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.axhline(1.0, color="#AAAAAA", lw=1.4, ls="--",
                   label=f"Referencia: {ref_impl.label}")
        for impl in cmp_impls:
            if impl not in avg_data:
                continue
            xs, ys = [], []
            for i, s in enumerate(sizes):
                t = avg_data[impl].get(s)
                r = ref_avgs.get(s)
                if t and r and t > 0:
                    xs.append(positions[i])
                    ys.append(r / t)
            if xs:
                ax.plot(xs, ys, marker="o", lw=2.2, color=impl.color, label=impl.label)
        _categorical_xticks(ax, sizes)
        ax.set_ylabel("Speedup  T(referencia) / T(implementación)")
        ax.set_xlabel("Tamaño de la carretera N")
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.legend(fontsize=8, loc="upper left")
        if note:
            ax.annotate(note, xy=(0.01, 0.03), xycoords="axes fraction",
                        fontsize=7, color="#888888", style="italic")
        fig.tight_layout()
    return _save_fig(fig, fname)


def chart_scaling(ref_impl: Impl, mpi_variants: list, avg_data: dict,
                  sizes: list, title: str, fname: str) -> str:
    """
    Escalabilidad fuerte: speedup y eficiencia vs número de procesos MPI,
    un panel por nivel de memoria (tamaño de carretera).
    """
    ref_avgs   = avg_data.get(ref_impl, {})
    np_counts  = sorted({i.np for i in mpi_variants if i.np >= 1})
    plot_sizes = sizes[:6]

    with plt.rc_context(CHART_STYLE):
        ncols = min(3, len(plot_sizes))
        nrows = math.ceil(len(plot_sizes) / ncols)
        fig, axs = plt.subplots(nrows, ncols,
                                figsize=(5 * ncols, 4 * nrows),
                                squeeze=False)

        for idx, size in enumerate(plot_sizes):
            ax    = axs[idx // ncols][idx % ncols]
            ref_t = ref_avgs.get(size)
            if not ref_t:
                ax.set_visible(False)
                continue

            sp_list, eff_list = [], []
            for p in np_counts:
                impl_here = next((i for i in mpi_variants if i.np == p), None)
                t = avg_data.get(impl_here, {}).get(size) if impl_here else None
                if t and t > 0:
                    sp = ref_t / t
                    sp_list.append(sp)
                    eff_list.append(sp / p * 100)
                else:
                    sp_list.append(float("nan"))
                    eff_list.append(float("nan"))

            ax2 = ax.twinx()
            ax2.spines["right"].set_visible(True)
            ax2.spines["right"].set_color("#70AD47")

            ax.plot(np_counts, sp_list,  "o-",  color="#2E75B6", lw=2,   label="Speedup")
            ax2.plot(np_counts, eff_list, "s--", color="#70AD47", lw=1.5, label="Eficiencia (%)")
            ax.plot(np_counts, [float(p) for p in np_counts],
                    "k--", lw=1, alpha=0.3, label="Escalado ideal")

            ax.set_title(f"N = {_size_label(size)}", fontsize=10)
            ax.set_xlabel("Número de procesos MPI")
            ax.set_ylabel("Speedup")
            ax2.set_ylabel("Eficiencia (%)", color="#70AD47")
            ax.set_xticks(np_counts)
            h1, l1 = ax.get_legend_handles_labels()
            h2, l2 = ax2.get_legend_handles_labels()
            ax.legend(h1 + h2, l1 + l2, fontsize=7, loc="upper left")

        for idx in range(len(plot_sizes), nrows * ncols):
            axs[idx // ncols][idx % ncols].set_visible(False)

        fig.suptitle(title, fontsize=12, fontweight="bold")
        fig.tight_layout()
    return _save_fig(fig, fname)


def chart_density_time(density_df: pd.DataFrame, best_seq: Impl,
                       mpi_opt_variants: list, title: str, fname: str) -> str:
    """Tiempo promedio vs densidad de tráfico para el experimento de densidad."""
    density_values = sorted(density_df["density"].unique())
    impls_to_plot  = [best_seq] + mpi_opt_variants

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 5))
        for impl in impls_to_plot:
            sub = density_df[
                (density_df["impl"] == impl.name) &
                (density_df["parallelism"] == impl.np)
            ]
            if sub.empty:
                continue
            agg = sub.groupby("density")["wall_time_ms"].mean().reindex(density_values)
            ax.plot(density_values, agg, marker="o", lw=2, color=impl.color, label=impl.label)
        ax.set_xlabel("Densidad de tráfico ρ")
        ax.set_ylabel("Tiempo promedio (ms)")
        ax.set_xticks(density_values)
        ax.legend(fontsize=8)
        ax.set_title(title, fontsize=11)
        fig.tight_layout()
    return _save_fig(fig, fname)


def chart_density_velocity(density_df: pd.DataFrame, best_seq: Impl,
                           mpi_opt_variants: list, title: str, fname: str) -> str:
    """Velocidad media del autómata vs densidad de tráfico."""
    density_values = sorted(density_df["density"].unique())
    impls_to_plot  = [best_seq] + mpi_opt_variants

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 5))
        for impl in impls_to_plot:
            sub = density_df[
                (density_df["impl"] == impl.name) &
                (density_df["parallelism"] == impl.np)
            ]
            if sub.empty:
                continue
            agg = sub.groupby("density")["avg_velocity"].mean().reindex(density_values)
            ax.plot(density_values, agg, marker="o", lw=2, color=impl.color, label=impl.label)

        # Curva teórica del modelo Nagel-Schreckenberg vmax=1
        rho_th = np.linspace(0.01, 0.99, 200)
        v_th   = np.where(rho_th <= 0.5, 1.0 - rho_th, (1.0 - rho_th) / rho_th)
        ax.plot(rho_th, v_th, "k--", lw=1.2, alpha=0.45, label="Velocidad teórica v∞(ρ)")

        ax.set_xlabel("Densidad de tráfico ρ")
        ax.set_ylabel("Velocidad media v̄")
        ax.set_xticks(density_values)
        ax.set_ylim(0, 1.05)
        ax.legend(fontsize=8)
        ax.set_title(title, fontsize=11)
        fig.tight_layout()
    return _save_fig(fig, fname)

# =============================================================================
# Tabla 1 — Mediciones individuales
# =============================================================================

def _write_raw_table(ws, impls: list, all_reps: dict,
                     sizes: list, n_cols: int, start_row: int) -> int:
    cur = start_row

    ws.merge_cells(f"A{cur}:{get_column_letter(n_cols)}{cur}")
    c = ws.cell(cur, 1, value="Tabla 1  —  Mediciones individuales (ms)")
    c.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur].height = 20
    cur += 1

    n_reps = max(
        (len(pairs) for impl in impls
         for pairs in all_reps.get(impl, {}).values()),
        default=1,
    )

    for impl_idx, impl in enumerate(impls):
        impl_reps = all_reps.get(impl, {})

        ws.merge_cells(f"A{cur}:{get_column_letter(n_cols)}{cur}")
        bh = ws.cell(cur, 1, value=impl.label)
        bh.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        bh.fill      = PatternFill("solid", fgColor=C["impl_hdr"])
        bh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        bh.border    = _thick_border()
        ws.row_dimensions[cur].height = 20
        cur += 1

        _hdr(ws.cell(cur, 1), "Rep.", bg=C["mid"], size=9)
        for ci, s in enumerate(sizes, 2):
            _hdr(ws.cell(cur, ci), f"N={_size_label(s)}", bg=C["mid"], size=9)
        ws.row_dimensions[cur].height = 18
        cur += 1

        for rep in range(1, n_reps + 1):
            bg = C["alt"] if rep % 2 == 0 else "FFFFFF"
            _dat(ws.cell(cur, 1), rep, bg=C["light"], bold=True, align="center")
            for ci, s in enumerate(sizes, 2):
                val = next((t for r, t, _ in impl_reps.get(s, []) if r == rep), None)
                if val is not None:
                    _dat(ws.cell(cur, ci), round(val, 3), fmt="#,##0.000", bg=bg)
                else:
                    _dat(ws.cell(cur, ci), "—", bg=bg, align="center")
            ws.row_dimensions[cur].height = 16
            cur += 1

        for s_label, s_idx in [("Promedio", 0), ("Desv. Est.", 1), ("CV (%)", 2)]:
            _hdr(ws.cell(cur, 1), s_label, bg=C["summary_bg"], fg=C["summary_fg"],
                 size=9, align="left", bold=True)
            for ci, s in enumerate(sizes, 2):
                vals = [t for _, t, _ in impl_reps.get(s, [])]
                if vals:
                    show = [_mean(vals), _std(vals), _cv(vals)][s_idx]
                    fmt  = ["#,##0.000", "#,##0.000", "0.00"][s_idx]
                    _summary_dat(ws.cell(cur, ci),
                                 round(show, 3 if s_idx < 2 else 2),
                                 fmt=fmt, bold=(s_idx == 0))
                else:
                    _summary_dat(ws.cell(cur, ci), "—")
            ws.row_dimensions[cur].height = 16
            cur += 1

        if impl_idx < len(impls) - 1:
            for ci in range(1, n_cols + 1):
                ws.cell(cur, ci).fill = PatternFill("solid", fgColor=C["sep"])
            ws.row_dimensions[cur].height = 6
            cur += 1

    return cur

# =============================================================================
# Tabla 2 — Promedios y speedup
# =============================================================================

def _write_speedup_table(ws, impls: list, avg_data: dict,
                         ref_impl: Impl, sizes: list, start_row: int) -> int:
    cur       = start_row
    n_sp_cols = 1 + len(sizes) * 2 + 1

    ws.merge_cells(f"A{cur}:{get_column_letter(n_sp_cols)}{cur}")
    c = ws.cell(cur, 1,
                value=f"Tabla 2  —  Tiempo promedio y Speedup  "
                      f"(referencia: {ref_impl.label})")
    c.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur].height = 20
    cur += 1

    _set_w(ws, 1, 38)
    for si in range(len(sizes)):
        _set_w(ws, 2 + si * 2, 13)
        _set_w(ws, 3 + si * 2, 11)
    _set_w(ws, n_sp_cols, 11)

    _hdr(ws.cell(cur, 1), "Implementación", bg=C["dark"], size=10)
    for si, s in enumerate(sizes):
        _hdr(ws.cell(cur, 2 + si * 2), f"N={_size_label(s)}\nPromedio (ms)", bg=C["mid"], size=9)
        _hdr(ws.cell(cur, 3 + si * 2), f"N={_size_label(s)}\nSpeedup",        bg=C["mid"], size=9)
    _hdr(ws.cell(cur, n_sp_cols), "Speedup\nPromedio", bg=C["dark"], size=9)
    ws.row_dimensions[cur].height = 30
    cur += 1

    ref_avgs = avg_data.get(ref_impl, {})
    for ri, impl in enumerate(impls):
        bg    = C["alt"] if ri % 2 == 0 else "FFFFFF"
        times = avg_data.get(impl, {})
        _dat(ws.cell(cur, 1), impl.label, bg=C["light"], bold=True, align="left")

        sp_cells = []
        for si, s in enumerate(sizes):
            avg_t   = times.get(s)
            ref_avg = ref_avgs.get(s)
            ac, sc  = 2 + si * 2, 3 + si * 2

            c_avg               = ws.cell(cur, ac)
            c_avg.value         = round(avg_t, 3) if avg_t else "N/D"
            c_avg.number_format = "#,##0.000"
            c_avg.font          = Font(name=FONT_NAME, size=10)
            c_avg.fill          = PatternFill("solid", fgColor=bg)
            c_avg.alignment     = Alignment(horizontal="right", vertical="center")
            c_avg.border        = _thin_border()

            c_sp = ws.cell(cur, sc)
            if impl == ref_impl:
                c_sp.value = 1.0
            elif avg_t and avg_t > 0 and ref_avg:
                c_sp.value = round(ref_avg / avg_t, 4)
                sp_cells.append(get_column_letter(sc) + str(cur))
            else:
                c_sp.value = "N/D"
            c_sp.number_format = "0.0000"
            c_sp.font          = Font(name=FONT_NAME, size=10, color=C["green_fg"])
            c_sp.fill          = PatternFill("solid", fgColor=C["green_bg"])
            c_sp.alignment     = Alignment(horizontal="right", vertical="center")
            c_sp.border        = _thin_border()

        c_av = ws.cell(cur, n_sp_cols)
        if sp_cells:
            c_av.value = f"=IFERROR(AVERAGE({','.join(sp_cells)}),\"N/D\")"
        elif impl == ref_impl:
            c_av.value = 1.0
        else:
            c_av.value = "N/D"
        c_av.number_format = "0.00"
        c_av.font          = Font(name=FONT_NAME, bold=True, size=10, color=C["green_fg"])
        c_av.fill          = PatternFill("solid", fgColor=C["green_bg"])
        c_av.alignment     = Alignment(horizontal="right", vertical="center")
        c_av.border        = _thin_border()
        ws.row_dimensions[cur].height = 17
        cur += 1

    return cur

# =============================================================================
# Escritor de hojas completas
# =============================================================================

def write_sheet(wb: Workbook, sheet_name: str, title: str,
                impls: list, all_reps: dict, avg_data: dict,
                ref_impl: Impl, sizes: list,
                chart_time_path: str, chart_sp_path: str,
                chart_scale_path: str | None = None):
    ws     = wb.create_sheet(sheet_name)
    n_cols = 1 + len(sizes)

    _title_row(ws, title, n_cols)
    _set_w(ws, 1, 38)
    for ci in range(2, n_cols + 1):
        _set_w(ws, ci, 16)

    cur = _write_raw_table(ws, impls, all_reps, sizes, n_cols, start_row=2)
    cur += 2
    cur = _write_speedup_table(ws, impls, avg_data, ref_impl, sizes, start_row=cur)
    cur += 3

    for anchor, path, w, h in [
        (f"A{cur}",      chart_time_path,  680, 390),
        (f"M{cur}",      chart_sp_path,    680, 390),
        (f"A{cur + 24}", chart_scale_path, 1050, 570),
    ]:
        if path and os.path.exists(path):
            img = XLImage(path)
            img.width = w
            img.height = h
            ws.add_image(img, anchor)

# =============================================================================
# Hoja de experimento de densidad
# =============================================================================

def write_density_sheet(wb: Workbook, density_df: pd.DataFrame,
                        best_seq: Impl, mpi_opt_variants: list,
                        chart_time_path: str, chart_vel_path: str):
    ws = wb.create_sheet("4. Densidad")
    density_values = sorted(density_df["density"].unique())
    n_cols         = len(density_values) + 1

    _title_row(ws, f"Experimento de densidad  —  N = {_size_label(DENSITY_EXPERIMENT_N)}", n_cols)
    _set_w(ws, 1, 38)
    for ci in range(2, n_cols + 1):
        _set_w(ws, ci, 14)

    cur          = 2
    impls_to_row = [best_seq] + mpi_opt_variants

    for metric_label, col, fmt in [
        ("Tiempo de ejecución promedio (ms)", "wall_time_ms", "#,##0.000"),
        ("Velocidad media del autómata v̄",   "avg_velocity", "0.000000"),
    ]:
        ws.merge_cells(f"A{cur}:{get_column_letter(n_cols)}{cur}")
        bh = ws.cell(cur, 1, value=metric_label)
        bh.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        bh.fill      = PatternFill("solid", fgColor=C["impl_hdr"])
        bh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[cur].height = 20
        cur += 1

        _hdr(ws.cell(cur, 1), "Implementación", bg=C["mid"], size=9)
        for ci, d in enumerate(density_values, 2):
            _hdr(ws.cell(cur, ci), f"ρ = {d:.2f}", bg=C["mid"], size=9)
        ws.row_dimensions[cur].height = 18
        cur += 1

        for ri, impl in enumerate(impls_to_row):
            sub = density_df[
                (density_df["impl"] == impl.name) &
                (density_df["parallelism"] == impl.np)
            ]
            bg = C["alt"] if ri % 2 == 0 else "FFFFFF"
            _dat(ws.cell(cur, 1), impl.label, bg=C["light"], bold=True, align="left")
            for ci, d in enumerate(density_values, 2):
                val = sub[sub["density"] == d][col].mean() if not sub.empty else math.nan
                _dat(ws.cell(cur, ci),
                     round(val, 4) if not math.isnan(val) else "—",
                     fmt=fmt, bg=bg)
            ws.row_dimensions[cur].height = 16
            cur += 1

        for ci in range(1, n_cols + 1):
            ws.cell(cur, ci).fill = PatternFill("solid", fgColor=C["sep"])
        ws.row_dimensions[cur].height = 8
        cur += 1

    cur += 2
    for anchor, path in [(f"A{cur}", chart_time_path), (f"M{cur}", chart_vel_path)]:
        if path and os.path.exists(path):
            img = XLImage(path)
            img.width = 680
            img.height = 400
            ws.add_image(img, anchor)
            
# =============================================================================
# Punto de entrada principal
# =============================================================================


def main():
    os.makedirs(CHARTS_DIR, exist_ok=True)

    df = load_benchmark_data(CSV_FILE_SEQ, CSV_FILE_MPI)
    if df.empty:
        print("Sin datos. Abortando.")
        sys.exit(1)

    scaling_df = get_scaling_df(df)
    density_df = get_density_df(df)
    all_reps   = build_all_reps(scaling_df)
    avg_data   = compute_avgs(all_reps)
    sizes      = sorted({s for szs in all_reps.values() for s in szs})

    # -----------------------------------------------------------------
    # Determinar referencias
    # -----------------------------------------------------------------
    best_seq = find_best_serial(avg_data, sizes)
    if best_seq is None:
        # Si solo hay MPI, usar una variante secuencial de fallback si existe.
        serials = [i for i in avg_data if i.np == 0]
        best_seq = serials[0] if serials else None
    if best_seq is None:
        print("No se encontraron variantes secuenciales en los datos.")
        sys.exit(1)

    mpi_variants     = get_mpi_variants(avg_data, NOMBRE_MPI)
    mpi_opt_variants = get_mpi_variants(avg_data, NOMBRE_MPI_OPT)
    best_mpi         = best_mpi_variant(NOMBRE_MPI,     avg_data, best_seq, sizes)
    best_mpi_opt     = best_mpi_variant(NOMBRE_MPI_OPT, avg_data, best_seq, sizes)

    serial_impls = [
        Impl(n, 0)
        for n in [NOMBRE_SEQ, NOMBRE_SEQ_OPT, NOMBRE_SEQ_MEM, NOMBRE_SEQ_MEM_OPT]
        if Impl(n, 0) in avg_data
    ]

    print(f"\n  Referencia secuencial óptima : {best_seq.label}")
    print(f"  Variantes MPI sin opt.       : {[i.label for i in mpi_variants]}")
    print(f"  Variantes MPI con opt.       : {[i.label for i in mpi_opt_variants]}")
    if best_mpi:
        print(f"  Mejor MPI sin opt.           : {best_mpi.label}")
    if best_mpi_opt:
        print(f"  Mejor MPI con opt.           : {best_mpi_opt.label}")

    # -----------------------------------------------------------------
    # Gráficas — Hoja 0: Secuenciales
    # -----------------------------------------------------------------
    print("\nGenerando gráficas...")

    ct0 = chart_time(
        serial_impls, avg_data, sizes,
        "Variantes secuenciales — tiempo de ejecución",
        "fig_seq_tiempo.png",
    )
    cs0 = chart_speedup(
        Impl(NOMBRE_SEQ, 0),
        [i for i in serial_impls if i != Impl(NOMBRE_SEQ, 0)],
        avg_data, sizes,
        "Variantes secuenciales — speedup vs Serial Base",
        "fig_seq_speedup.png",
    ) if Impl(NOMBRE_SEQ, 0) in avg_data else chart_speedup(
        best_seq,
        [i for i in serial_impls if i != best_seq],
        avg_data, sizes,
        "Variantes secuenciales — speedup vs referencia",
        "fig_seq_speedup.png",
    )

    # -----------------------------------------------------------------
    # Gráficas — Hoja 1: MPI sin optimización
    # -----------------------------------------------------------------
    ct1 = chart_time(
        [best_seq] + mpi_variants, avg_data, sizes,
        "MPI sin optimización de compilador — tiempo de ejecución",
        "fig_mpi_tiempo.png",
    )
    cs1 = chart_speedup(
        best_seq, mpi_variants, avg_data, sizes,
        "MPI sin optimización de compilador — speedup",
        "fig_mpi_speedup.png",
        note="Speedup esperado ≤ número de procesos (escalado ideal)",
    )
    csc1 = chart_scaling(
        best_seq, mpi_variants, avg_data, sizes,
        "MPI sin optimización — escalabilidad fuerte por nivel de memoria",
        "fig_mpi_escalabilidad.png",
    ) if mpi_variants else None

    # -----------------------------------------------------------------
    # Gráficas — Hoja 2: MPI con optimización
    # -----------------------------------------------------------------
    ct2 = chart_time(
        [best_seq] + mpi_opt_variants, avg_data, sizes,
        "MPI con optimización de compilador — tiempo de ejecución",
        "fig_mpi_opt_tiempo.png",
    )
    cs2 = chart_speedup(
        best_seq, mpi_opt_variants, avg_data, sizes,
        "MPI con optimización de compilador — speedup",
        "fig_mpi_opt_speedup.png",
    )
    csc2 = chart_scaling(
        best_seq, mpi_opt_variants, avg_data, sizes,
        "MPI con optimización — escalabilidad fuerte por nivel de memoria",
        "fig_mpi_opt_escalabilidad.png",
    ) if mpi_opt_variants else None

    # -----------------------------------------------------------------
    # Gráficas — Hoja 3: Comparación final
    # -----------------------------------------------------------------
    final_impls = [i for i in [best_seq, best_mpi, best_mpi_opt] if i]
    ct3 = chart_time(
        final_impls, avg_data, sizes,
        "Comparación final — mejor de cada estrategia",
        "fig_final_tiempo.png",
    )
    cs3 = chart_speedup(
        best_seq,
        [i for i in final_impls if i != best_seq],
        avg_data, sizes,
        "Comparación final — speedup de los mejores vs referencia secuencial",
        "fig_final_speedup.png",
    )

    # -----------------------------------------------------------------
    # Gráficas — Hoja 4: Densidad
    # -----------------------------------------------------------------
    c_dt = c_dv = None
    if not density_df.empty:
        density_reps   = build_all_reps(density_df)
        density_avgs   = compute_avgs(density_reps)
        mpi_opt_den    = get_mpi_variants(density_avgs, NOMBRE_MPI_OPT)
        best_seq_den   = (
            best_seq if Impl(best_seq.name, 0) in density_avgs
            else (get_mpi_variants(density_avgs, best_seq.name) or [best_seq])[0]
        )
        c_dt = chart_density_time(
            density_df, best_seq_den, mpi_opt_den,
            f"Densidad — tiempo de ejecución  (N = {_size_label(DENSITY_EXPERIMENT_N)})",
            "fig_densidad_tiempo.png",
        )
        c_dv = chart_density_velocity(
            density_df, best_seq_den, mpi_opt_den,
            f"Densidad — velocidad media del autómata  (N = {_size_label(DENSITY_EXPERIMENT_N)})",
            "fig_densidad_velocidad.png",
        )

    # -----------------------------------------------------------------
    # Construcción del libro Excel
    # -----------------------------------------------------------------
    print("\nConstruyendo libro de Excel...")
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    write_sheet(
        wb, "0. Secuenciales",
        "Etapa 0 — Comparación entre variantes secuenciales",
        serial_impls if serial_impls else [best_seq],
        all_reps, avg_data,
        Impl(NOMBRE_SEQ, 0) if Impl(NOMBRE_SEQ, 0) in avg_data else best_seq,
        sizes, ct0, cs0,
    )

    write_sheet(
        wb, "1. MPI sin optimización",
        f"Etapa 1 — {best_seq.label} vs MPI sin optimización de compilador",
        [best_seq] + mpi_variants,
        all_reps, avg_data,
        best_seq, sizes, ct1, cs1, csc1,
    )

    write_sheet(
        wb, "2. MPI con optimización",
        f"Etapa 2 — {best_seq.label} vs MPI con optimización de compilador",
        [best_seq] + mpi_opt_variants,
        all_reps, avg_data,
        best_seq, sizes, ct2, cs2, csc2,
    )

    if final_impls:
        write_sheet(
            wb, "3. Comparación final",
            "Etapa 3 — Mejor variante secuencial vs mejores variantes MPI",
            final_impls,
            all_reps, avg_data,
            best_seq, sizes, ct3, cs3,
        )

    if not density_df.empty:
        density_reps_for_sheet = build_all_reps(density_df)
        mpi_opt_den_for_sheet  = get_mpi_variants(
            compute_avgs(density_reps_for_sheet), NOMBRE_MPI_OPT
        )
        best_seq_in_den = best_seq
        write_density_sheet(
            wb, density_df,
            best_seq_in_den, mpi_opt_den_for_sheet,
            c_dt, c_dv,
        )

    wb.save(OUTPUT_PATH)
    print(f"\n  Listo.")
    print(f"  Reporte : {OUTPUT_PATH}")
    print(f"  Gráficas: {CHARTS_DIR}/")


if __name__ == "__main__":
    main()