"""
report_profiling.py — Traffic Automaton Sequential Profiling Report
===================================================================

Versión enfocada en:
  - hotspots reales del kernel
  - detalle interno de funciones a partir de gprof y perf
  - menor interés en branch prediction
  - gráficas más útiles para análisis del loop principal

Reads:
  - data_profiling.csv
  - raw/N<size>/gprof_report.txt
  - raw/N<size>/perf_stat.txt
  - raw/N<size>/perf_record_report.txt
  - raw/N<size>/cachegrind_report.txt
  - raw/N<size>/massif_report.txt
  - raw/N<size>/timing_ms.txt
  - raw/N<size>/timing_vel.txt

Writes:
  - reporte_profiling.xlsx
  - profiling_report.html
  - profiling_table.tex
  - charts_profiling/*.png

Usage:
  python3 report_profiling.py [path/to/data_profiling.csv]
"""

from __future__ import annotations

import base64
import html
import math
import os
import re
import sys
import statistics
from typing import Any
from pprint import pprint
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


CSV_PATH = (
    sys.argv[1]
    if len(sys.argv) > 1
    else "tests/machine1/results_profiling/data_profiling.csv"
)

OUT_DIR    = os.path.dirname(os.path.abspath(CSV_PATH))
RAW_DIR    = os.path.join(OUT_DIR, "raw")
CHARTS_DIR = os.path.join(OUT_DIR, "charts_profiling")
XLSX_PATH  = os.path.join(OUT_DIR, "reporte_profiling.xlsx")
HTML_PATH  = os.path.join(OUT_DIR, "profiling_report.html")
LATEX_PATH = os.path.join(OUT_DIR, "profiling_table.tex")


# ---------------------------------------------------------------------------
# Style
# ---------------------------------------------------------------------------
COLORS = {
    "navy":     "1F3557",
    "blue":     "2E75B6",
    "cyan":     "1F9EB7",
    "green":    "70AD47",
    "orange":   "ED7D31",
    "red":      "C00000",
    "purple":   "7030A0",
    "gray":     "808080",
    "light":    "DCE6F1",
    "lighter":  "F7FBFF",
    "alt":      "F4F8FC",
    "white":    "FFFFFF",
    "border":   "C9D6E2",
    "good_bg":  "E2F0D9",
    "good_fg":  "2F6B1E",
    "warn_bg":  "FFF2CC",
    "warn_fg":  "7F6000",
    "bad_bg":   "FCE4D6",
    "bad_fg":   "A61C00",
}

CHART_STYLE = {
    "figure.facecolor":    "white",
    "axes.facecolor":      "white",
    "axes.grid":           True,
    "grid.color":          "#E7EDF3",
    "grid.linestyle":      "--",
    "grid.alpha":          0.75,
    "axes.spines.top":     False,
    "axes.spines.right":   False,
    "axes.titleweight":    "bold",
    "axes.labelsize":      10,
    "axes.titlesize":      12,
    "legend.frameon":      False,
    "font.size":           10,
}
plt.rcParams.update(CHART_STYLE)


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
def _border() -> Border:
    s = Side(style="thin", color=COLORS["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(key: str) -> PatternFill:
    return PatternFill("solid", fgColor=COLORS[key])

def _set_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width

def _title_row(ws, text: str, cols: int, row: int = 1) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Calibri", bold=True, size=14, color=COLORS["white"])
    c.fill      = _fill("navy")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _border()
    ws.row_dimensions[row].height = 24

def _hcell(cell, value: Any, bg: str = "blue") -> None:
    cell.value     = value
    cell.font      = Font(name="Calibri", bold=True, size=10, color=COLORS["white"])
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border()

def _dcell(cell, value: Any, fmt: str | None = None, bg: str = "white",
           bold: bool = False, align: str = "right", fg: str = "000000") -> None:
    cell.value     = value
    cell.font      = Font(name="Calibri", bold=bold, size=10, color=COLORS.get(fg, fg))
    cell.fill      = _fill(bg if bg in COLORS else "white")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = _border()
    if fmt:
        cell.number_format = fmt


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------
def _canon(name: str) -> str:
    s = name.strip().lower().replace("%", "pct").replace("-", "_").replace(" ", "_")
    s = re.sub(r"[^a-z0-9_]", "", s)
    return re.sub(r"_+", "_", s).strip("_")

def _read(path: str) -> str:
    if not os.path.exists(path):
        return ""
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def _sf(x: Any, default: float = math.nan) -> float:
    try:
        return math.nan if pd.isna(x) else float(x)
    except Exception:
        return default

def _si(x: Any, default: int = 0) -> int:
    try:
        return default if pd.isna(x) else int(float(x))
    except Exception:
        return default

def _fn(x: Any, d: int = 3) -> str:
    try:
        f = float(x)
        return "—" if math.isnan(f) else f"{f:,.{d}f}"
    except Exception:
        return "—"

def _fp(x: Any, d: int = 2) -> str:
    try:
        f = float(x)
        return "—" if math.isnan(f) else f"{f:.{d}f}%"
    except Exception:
        return "—"

def _fi(x: Any) -> str:
    try:
        f = float(x)
        return "—" if math.isnan(f) else f"{int(round(f)):,}"
    except Exception:
        return "—"

def _classify_ipc(v: float) -> str:
    if math.isnan(v): return "Sin dato"
    if v >= 2.5:      return "Alto"
    if v >= 1.5:      return "Medio"
    return "Bajo"

def _classify_miss(v: float) -> str:
    if math.isnan(v): return "Sin dato"
    if v < 1:         return "Excelente"
    if v < 5:         return "Bueno"
    if v < 15:        return "Moderado"
    return "Alto"

def theoretical_heap_mb(n: int) -> float:
    return 2 * n * 4 / 1024 / 1024

def perf_extract(path: str, keyword: str) -> int:
    if not os.path.exists(path):
        return 0
    pattern = re.compile(
        r"^\s*[\d,]+.*\s" + re.escape(keyword) + r"(\s|$)",
        re.IGNORECASE,
    )
    for line in _read(path).splitlines():
        if line.lstrip().startswith("#"):
            continue
        if pattern.search(line):
            m = re.match(r"^\s*([\d,]+)", line)
            if m:
                return int(m.group(1).replace(",", ""))
    return 0

def compute_ratio(num: int, den: int, decimals: int = 4) -> float:
    return round(num / den * 100, decimals) if den > 0 else 0.0

def save_figure(fig, name: str) -> str:
    path = os.path.join(CHARTS_DIR, name)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return path

def annotate_points(ax, xs, ys, d: int = 3, dy: int = 8):
    for x, y in zip(xs, ys):
        if pd.isna(y):
            continue
        ax.annotate(f"{y:.{d}f}", (x, y), textcoords="offset points",
                    xytext=(0, dy), ha="center", fontsize=8)

def _cat_axis(sizes: list[int]) -> tuple[list[int], list[str]]:
    positions = list(range(len(sizes)))
    labels    = [f"N={n:,}" for n in sizes]
    return positions, labels

def _set_cat_xticks(ax, positions, labels, rotation: int = 20) -> None:
    ax.set_xticks(positions)
    ax.set_xticklabels(labels, rotation=rotation, ha="right")
    ax.set_xlabel("N (road length)")

def _clean_symbol_name(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+\(.*\)$", "", s)
    s = re.sub(r"\+0x[0-9a-fA-F]+$", "", s)
    return s.strip()

def _shorten(s: str, n: int = 42) -> str:
    s = str(s)
    return s if len(s) <= n else s[:n-1] + "…"

def _best_hot_name(raw_entry: dict[str, Any]) -> str:
    gtop = raw_entry.get("gprof", {}).get("top", [])
    if gtop:
        return gtop[0]["name"]
    ptop = raw_entry.get("perf_record", {}).get("top", [])
    if ptop:
        return _clean_symbol_name(ptop[0]["symbol"])
    return ""


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def normalize_dataframe(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = [_canon(c) for c in df.columns]

    rename_map = {"time_ms": "time_mean_ms"}
    df = df.rename(columns={c: rename_map.get(c, c) for c in df.columns})

    if "road_length" not in df.columns:
        raise ValueError("CSV inválido: falta la columna road_length")

    for col in df.columns:
        if col == "road_length":
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
        else:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.sort_values("road_length").reset_index(drop=True)

    for col in ("time_mean_ms", "time_std_ms", "avg_velocity",
                "throughput_mcells_s", "ipc", "peak_heap_mb"):
        if col not in df.columns:
            df[col] = math.nan

    if "peak_heap_mb" in df.columns:
        df["peak_heap_mb"] = df["peak_heap_mb"].where(df["peak_heap_mb"] > 0, other=np.nan)

    df["theoretical_heap_mb"] = df["road_length"].apply(
        lambda n: theoretical_heap_mb(int(n)) if pd.notna(n) else math.nan
    )
    df["heap_ratio_pct"] = 100.0 * df["peak_heap_mb"] / df["theoretical_heap_mb"]
    df["timing_cv_pct"]  = 100.0 * df["time_std_ms"] / df["time_mean_ms"]

    return df


# ---------------------------------------------------------------------------
# Raw report parsers
# ---------------------------------------------------------------------------
def parse_timing_traffic(raw_dir: str) -> dict[str, Any]:
    ms_vals, vel_vals = [], []

    for line in _read(os.path.join(raw_dir, "timing_ms.txt")).splitlines():
        line = line.strip()
        if line:
            try:
                ms_vals.append(float(line))
            except ValueError:
                pass

    for line in _read(os.path.join(raw_dir, "timing_vel.txt")).splitlines():
        line = line.strip()
        if line:
            try:
                vel_vals.append(float(line))
            except ValueError:
                pass

    if not ms_vals:
        return {
            "values": [], "vel_values": [], "count": 0,
            "mean_ms": math.nan, "std_ms": math.nan,
            "min_ms": math.nan, "max_ms": math.nan,
            "median_ms": math.nan, "cv_pct": math.nan,
            "mean_vel": math.nan, "std_vel": math.nan
        }

    mean     = statistics.mean(ms_vals)
    std      = statistics.stdev(ms_vals) if len(ms_vals) > 1 else 0.0
    mean_vel = statistics.mean(vel_vals) if vel_vals else math.nan
    std_vel  = statistics.stdev(vel_vals) if len(vel_vals) > 1 else 0.0

    return {
        "values": ms_vals,
        "vel_values": vel_vals,
        "count": len(ms_vals),
        "mean_ms": mean,
        "std_ms": std,
        "min_ms": min(ms_vals),
        "max_ms": max(ms_vals),
        "median_ms": statistics.median(ms_vals),
        "cv_pct": 100.0 * std / mean if mean else math.nan,
        "mean_vel": mean_vel,
        "std_vel": std_vel,
    }
    
def parse_perf_mem(path: str, top_k: int = 12) -> dict[str, Any]:
    text = _read(path)
    if not text:
        return {"total_samples": 0, "levels": {}, "symbols": [],
                "symbols_by_level": {}, "na_pct": 0.0}

    # Cabecera
    total_samples = 0
    m = re.search(r"Samples:\s+([\d.]+)K?\s+of event", text)
    if m:
        val = m.group(1)
        total_samples = int(float(val) * 1000) if "K" in text[m.start():m.end()+5] else int(val)

    # Columnas reales del sort order:
    # Overhead | Samples | Local Weight | Memory access | Symbol | Shared Object |
    # Data Symbol | Data Object | Snoop | TLB access | Locked | Blocked |
    # Local INSTR Latency | Local Retire Latency

    levels: dict[str, float] = {}        # nivel → % overhead acumulado
    symbols: list[dict] = []             # por símbolo, acumulado
    sym_level: dict[str, dict] = {}      # símbolo → {nivel → pct}
    na_overhead = 0.0

    sym_acc: dict[str, float] = {}

    line_re = re.compile(
        r"^\s*([\d.]+)%\s+"          # overhead
        r"(\d+)\s+"                   # samples
        r"(\d+)\s+"                   # local weight
        r"(.+?)\s{2,}"                # memory access (campo de ancho variable)
        r"\[\.\]\s+(\S+)"             # symbol (después de "[.] ")
    )

    # alternativa para kernel [k]
    line_re_k = re.compile(
        r"^\s*([\d.]+)%\s+"
        r"(\d+)\s+"
        r"(\d+)\s+"
        r"(.+?)\s{2,}"
        r"\[k\]\s+(\S+)"
    )

    for line in text.splitlines():
        if line.lstrip().startswith("#") or not line.strip():
            continue

        m = line_re.search(line) or line_re_k.search(line)
        if not m:
            continue

        try:
            overhead = float(m.group(1))
            mem_level = m.group(4).strip()
            symbol    = _clean_symbol_name(m.group(5))
        except (ValueError, IndexError):
            continue

        # Acumular por nivel
        if mem_level == "N/A":
            na_overhead += overhead
        else:
            levels[mem_level] = levels.get(mem_level, 0.0) + overhead

        # Acumular por símbolo (solo user-space del binario)
        if "[.] " in line and "traffic_seq" in line:
            sym_acc[symbol] = sym_acc.get(symbol, 0.0) + overhead
            if symbol not in sym_level:
                sym_level[symbol] = {}
            if mem_level != "N/A":
                sym_level[symbol][mem_level] = sym_level[symbol].get(mem_level, 0.0) + overhead

    symbols = [
        {"symbol": sym, "overhead_pct": pct,
         "levels": sym_level.get(sym, {})}
        for sym, pct in sorted(sym_acc.items(), key=lambda x: x[1], reverse=True)
    ]

    return {
        "total_samples": total_samples,
        "levels": levels,
        "na_pct": na_overhead,
        "symbols": symbols[:top_k],
        "symbols_by_level": sym_level,
    }

def parse_gprof_flat(path: str, top_k: int = 10) -> dict[str, Any]:
    text = _read(path)
    if not text:
        return {"sample_seconds": math.nan, "top": []}

    m = re.search(r"Each sample counts as\s+([\d.]+)\s+seconds", text)
    sample_seconds = float(m.group(1)) if m else math.nan

    rows = []
    in_flat = False
    for line in text.splitlines():
        if line.strip().startswith("Flat profile"):
            in_flat = True
            continue
        if not in_flat:
            continue
        if line.strip().startswith("%"):
            continue
        if line.startswith("\f"):
            break
        if not line.strip():
            continue

        parts = line.split()
        if len(parts) < 7:
            continue
        try:
            pct_time     = float(parts[0])
            cumulative   = float(parts[1])
            self_seconds = float(parts[2])
        except ValueError:
            continue

        calls = int(parts[3]) if parts[3].isdigit() else None
        name = parts[-1]
        rows.append({
            "name": name,
            "pct_time": pct_time,
            "cumulative_s": cumulative,
            "self_s": self_seconds,
            "calls": calls
        })

    return {"sample_seconds": sample_seconds, "top": rows[:top_k]}


def parse_gprof_call_graph(path: str, top_k: int = 12) -> dict[str, Any]:
    text = _read(path)
    if not text:
        return {"functions": [], "edges": []}

    blocks = []
    current = []
    in_call_graph = False

    for line in text.splitlines():
        if "Call graph" in line:
            in_call_graph = True
            current = []
            continue
        if not in_call_graph:
            continue

        if re.match(r"^-{10,}$", line.strip()):
            if current:
                blocks.append("\n".join(current))
                current = []
            continue

        current.append(line)

    if current:
        blocks.append("\n".join(current))

    functions = []
    edges = []

    entry_re = re.compile(
        r"^\[(\d+)\]\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+(\d+)?/?(\d+)?\s+(.+)$"
    )

    caller_re = re.compile(
        r"^\s+([\d.]+)\s+([\d.]+)\s+(\d+)?/?(\d+)?\s+(.+)\s+\[(\d+)\]$"
    )

    callee_re = re.compile(
        r"^\s+([\d.]+)\s+([\d.]+)\s+(\d+)?/?(\d+)?\s+(.+)\s+\[(\d+)\]$"
    )

    for blk in blocks:
        lines = [ln.rstrip() for ln in blk.splitlines() if ln.strip()]
        center = None

        for i, line in enumerate(lines):
            m = entry_re.match(line.strip())
            if m:
                center = {
                    "index": int(m.group(1)),
                    "pct_time": float(m.group(2)),
                    "self_s": float(m.group(3)),
                    "children_s": float(m.group(4)),
                    "called_self": int(m.group(5)) if m.group(5) else None,
                    "called_total": int(m.group(6)) if m.group(6) else None,
                    "name": _clean_symbol_name(m.group(7)),
                    "raw": line.strip(),
                    "parents": [],
                    "children": [],
                }
                upper = lines[:i]
                lower = lines[i+1:]

                for up in upper:
                    mu = caller_re.match(up)
                    if mu:
                        center["parents"].append({
                            "self_s": float(mu.group(1)),
                            "children_s": float(mu.group(2)),
                            "called_self": int(mu.group(3)) if mu.group(3) else None,
                            "called_total": int(mu.group(4)) if mu.group(4) else None,
                            "name": _clean_symbol_name(mu.group(5)),
                            "index": int(mu.group(6)),
                        })

                for lo in lower:
                    ml = callee_re.match(lo)
                    if ml:
                        center["children"].append({
                            "self_s": float(ml.group(1)),
                            "children_s": float(ml.group(2)),
                            "called_self": int(ml.group(3)) if ml.group(3) else None,
                            "called_total": int(ml.group(4)) if ml.group(4) else None,
                            "name": _clean_symbol_name(ml.group(5)),
                            "index": int(ml.group(6)),
                        })
                break

        if center:
            functions.append(center)
            for ch in center["children"]:
                edges.append({
                    "src": center["name"],
                    "dst": ch["name"],
                    "time_s": ch["self_s"] + ch["children_s"],
                    "calls": ch["called_total"],
                })

    functions = sorted(functions, key=lambda x: x["pct_time"], reverse=True)
    return {"functions": functions[:top_k], "edges": edges}


def parse_perf_record(path: str, top_k: int = 12) -> dict[str, Any]:
    text = _read(path)
    if not text:
        return {"samples": 0, "event_count": 0, "top": []}

    m = re.search(r"Samples:\s+(\d+)", text)
    samples = int(m.group(1)) if m else 0
    m = re.search(r"Event count \(approx\.\):\s+([\d,]+)", text)
    event_count = int(m.group(1).replace(",", "")) if m else 0

    rows = []
    for line in text.splitlines():
        if line.lstrip().startswith("#") or not line.strip():
            continue
        m = re.match(r"^\s*([\d.]+)%\s+(\S+)\s+(\S+)\s+(\[[^\]]+\])\s+(.+?)\s*$", line)
        if m:
            rows.append({
                "overhead_pct": float(m.group(1)),
                "command": m.group(2),
                "shared_object": m.group(3),
                "symbol": _clean_symbol_name(m.group(5)),
            })

    return {"samples": samples, "event_count": event_count, "top": rows[:top_k]}


def parse_perf_record_hierarchy(path: str, top_k: int = 12) -> dict[str, Any]:
    text = _read(path)
    if not text:
        return {"groups": []}

    rows = []
    for line in text.splitlines():
        if line.lstrip().startswith("#") or not line.strip():
            continue
        m = re.match(r"^\s*([\d.]+)%\s+(\S+)\s+(\S+)\s+(\[[^\]]+\])\s+(.+?)\s*$", line)
        if m:
            rows.append({
                "overhead_pct": float(m.group(1)),
                "command": m.group(2),
                "shared_object": m.group(3),
                "symbol": _clean_symbol_name(m.group(5)),
            })

    if not rows:
        return {"groups": []}

    groups = {}
    for r in rows:
        key = r["symbol"]
        if key not in groups:
            groups[key] = {
                "symbol": key,
                "overhead_pct": 0.0,
                "objects": {},
            }
        groups[key]["overhead_pct"] += r["overhead_pct"]
        so = r["shared_object"]
        groups[key]["objects"][so] = groups[key]["objects"].get(so, 0.0) + r["overhead_pct"]

    grouped = []
    for sym, info in groups.items():
        grouped.append({
            "symbol": sym,
            "overhead_pct": info["overhead_pct"],
            "top_object": max(info["objects"].items(), key=lambda x: x[1])[0]
        })

    grouped = sorted(grouped, key=lambda x: x["overhead_pct"], reverse=True)
    return {"groups": grouped[:top_k]}


def parse_cachegrind(path: str, top_k: int = 10, top_lines: int = 10) -> dict[str, Any]:
    text = _read(path)
    if not text:
        return {"total_ir": 0, "top_functions": [], "hot_lines": []}

    total_ir = 0
    m = re.search(r"^\s*([\d,]+)\s+\(100\.0%\)\s+PROGRAM TOTALS", text, re.MULTILINE)
    if m:
        total_ir = int(m.group(1).replace(",", ""))

    top_functions, in_func = [], False
    for line in text.splitlines():
        if "-- Function:file summary" in line:
            in_func = True
            continue
        if in_func and line.startswith("-- Annotated source file"):
            break
        if not in_func:
            continue

        m = re.match(r"^>\s*([\d,]+)\s+\(([\d.]+)%.*?\)\s+(.+?)\s*$", line.strip())
        if not m:
            m = re.match(r"^\s*([\d,]+)\s+\(([\d.]+)%[^)]*\)\s+(.+?)\s*$", line.strip())
        if m:
            top_functions.append({
                "ir": int(m.group(1).replace(",", "")),
                "pct": float(m.group(2)),
                "name": _clean_symbol_name(m.group(3))
            })

    hot_lines = []
    for line in text.splitlines():
        m = re.match(r"^\s*([\d,]+)\s+\(([\d.]+)%\)\s+(.+?)\s*$", line)
        if not m:
            continue
        code = m.group(3).strip()
        if not code or code == "." or code.startswith("--") or code.startswith("Unannotated"):
            continue
        hot_lines.append({
            "ir": int(m.group(1).replace(",", "")),
            "pct": float(m.group(2)),
            "code": code
        })

    hot_lines = sorted(hot_lines, key=lambda x: x["ir"], reverse=True)[:top_lines]
    top_functions = sorted(top_functions, key=lambda x: x["ir"], reverse=True)[:top_k]

    return {
        "total_ir": total_ir,
        "top_functions": top_functions,
        "hot_lines": hot_lines
    }


def parse_massif(path: str) -> dict[str, Any]:
    text = _read(path)
    empty = {
        "snapshots": 0,
        "peak_total_b": 0,
        "peak_useful_b": 0,
        "peak_extra_b": 0,
        "peak_stacks_b": 0,
        "peak_total_mb": math.nan,
        "peak_useful_pct": math.nan,
        "allocators": []
    }

    if not text or text.strip() == "SKIPPED":
        return empty

    m = re.search(r"Number of snapshots:\s+(\d+)", text)
    snapshots = int(m.group(1)) if m else 0

    rows = []
    for line in text.splitlines():
        m = re.match(r"^\s*(\d+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s*$", line)
        if m:
            rows.append({
                "total_b": int(m.group(3).replace(",", "")),
                "useful_b": int(m.group(4).replace(",", "")),
                "extra_b": int(m.group(5).replace(",", "")),
                "stacks_b": int(m.group(6).replace(",", "")),
            })

    if not rows:
        return {**empty, "snapshots": snapshots}

    peak = max(rows, key=lambda x: x["total_b"])
    peak_mb = peak["total_b"] / 1024 / 1024
    peak_pct = 100.0 * peak["useful_b"] / peak["total_b"] if peak["total_b"] else math.nan

    return {
        "snapshots": snapshots,
        "peak_total_b": peak["total_b"],
        "peak_useful_b": peak["useful_b"],
        "peak_extra_b": peak["extra_b"],
        "peak_stacks_b": peak["stacks_b"],
        "peak_total_mb": peak_mb,
        "peak_useful_pct": peak_pct,
        "allocators": []
    }


def collect_raw(df: pd.DataFrame) -> dict[int, dict[str, Any]]:
    results: dict[int, dict[str, Any]] = {}
    for n in df["road_length"].dropna().astype(int).tolist():
        d = os.path.join(RAW_DIR, f"N{n}")
        results[n] = {
            "timing":        parse_timing_traffic(d),
            "gprof":         parse_gprof_flat(os.path.join(d, "gprof_report.txt")),
            "gprof_call":    parse_gprof_call_graph(os.path.join(d, "gprof_report.txt")),
            "perf_record":   parse_perf_record(os.path.join(d, "perf_record_report.txt")),
            "perf_mem":      parse_perf_mem(os.path.join(d, "perf_mem_report.txt")),
            "perf_hier":     parse_perf_record_hierarchy(os.path.join(d, "perf_record_report.txt")),
            "cachegrind":    parse_cachegrind(os.path.join(d, "cachegrind_report.txt")),
            "massif":        parse_massif(os.path.join(d, "massif_report.txt")),
            "perf_stat_txt": _read(os.path.join(d, "perf_stat.txt")),
        }
    return results


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------
def chart_performance(df: pd.DataFrame) -> str:
    sizes = df["road_length"].astype(int).tolist()
    pos, labels = _cat_axis(sizes)

    fig, axs = plt.subplots(2, 2, figsize=(13, 8))

    axs[0, 0].errorbar(pos, df["time_mean_ms"], yerr=df["time_std_ms"],
                       marker="o", lw=2.4, capsize=4, color="#2E75B6")
    axs[0, 0].set_title("Tiempo medio de ejecución")
    axs[0, 0].set_ylabel("ms")

    if "throughput_mcells_s" in df.columns and df["throughput_mcells_s"].notna().any():
        tvals = df["throughput_mcells_s"].dropna()
        axs[0, 1].plot(pos, df["throughput_mcells_s"], marker="o", lw=2.4, color="#1F9EB7")
        annotate_points(axs[0, 1], pos, df["throughput_mcells_s"], d=2)
        axs[0, 1].set_title("Throughput secuencial")
        axs[0, 1].set_ylabel("Mcells/s")
        y_lo = max(0.0, float(tvals.min()) * 0.995)
        y_hi = float(tvals.max()) * 1.005
        axs[0, 1].set_ylim(y_lo, y_hi)

    axs[1, 0].plot(pos, df["ipc"], marker="D", lw=2.4, color="#7030A0")
    axs[1, 0].axhline(1.0, color="#BBBBBB", lw=1, ls="--")
    axs[1, 0].axhline(2.0, color="#BBBBBB", lw=1, ls="--")
    annotate_points(axs[1, 0], pos, df["ipc"])
    axs[1, 0].set_title("Eficiencia del pipeline (IPC)")
    axs[1, 0].set_ylabel("IPC")

    heap_measured = df["peak_heap_mb"].where(df["peak_heap_mb"] > 0, other=np.nan)
    axs[1, 1].plot(pos, heap_measured, marker="o", lw=2.4,
                   color="#70AD47", label="Massif (medido)")
    axs[1, 1].plot(pos, df["theoretical_heap_mb"], marker="s", lw=2.0,
                   ls="--", color="#7F7F7F", label="Teórico (2×N×4 B)")
    axs[1, 1].set_title("Memoria pico heap")
    axs[1, 1].set_ylabel("MB")
    axs[1, 1].legend()

    for ax in axs.flat:
        _set_cat_xticks(ax, pos, labels)

    fig.suptitle("Resumen de rendimiento y memoria — autómata secuencial",
                 fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "01_performance_summary.png")


def chart_velocity(df: pd.DataFrame) -> str:
    sizes = df["road_length"].astype(int).tolist()
    pos, labels = _cat_axis(sizes)

    fig, ax = plt.subplots(figsize=(9, 4.5))
    ax.plot(pos, df["avg_velocity"], marker="o", lw=2.4, color="#ED7D31")
    ax.fill_between(pos, df["avg_velocity"], alpha=0.12, color="#ED7D31")
    annotate_points(ax, pos, df["avg_velocity"], d=4)
    ax.set_title("Velocidad media asintótica vs tamaño de carretera (ρ=0.50)")
    ax.set_ylabel("v̄")
    ax.set_ylim(0, 1.05)
    ax.axhline(0.5, color="#BBBBBB", lw=1, ls="--")
    _set_cat_xticks(ax, pos, labels)
    fig.tight_layout()
    return save_figure(fig, "02_avg_velocity.png")

def chart_perf_mem(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> str:
    largest_n = int(df["road_length"].max())
    mem_data  = raw.get(largest_n, {}).get("perf_mem", {})
    levels    = mem_data.get("levels", {})
    symbols   = mem_data.get("symbols", [])
    na_pct    = mem_data.get("na_pct", 0.0)

    fig, axs = plt.subplots(1, 2, figsize=(14, 5.8))

    # Panel izquierdo: distribución global por nivel
    if levels:
        all_levels = dict(sorted(levels.items(), key=lambda x: x[1], reverse=True))
        if na_pct > 0:
            all_levels["N/A (IBS sin resolución)"] = na_pct

        lnames = list(all_levels.keys())[::-1]
        lvals  = [all_levels[k] for k in lnames[::-1]]
        palette = ["#70AD47","#1F9EB7","#2E75B6","#7030A0","#C00000","#BBBBBB"]
        bars = axs[0].barh(lnames, lvals,
                           color=palette[:len(lnames)])
        mx = max(lvals) if lvals else 1
        for bar, val in zip(bars, lvals):
            axs[0].text(
                bar.get_width() + mx * 0.02,
                bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}%", va="center", fontsize=9
            )
        axs[0].set_title(f"Accesos resueltos por nivel (N={largest_n:,})")
        axs[0].set_xlabel("% overhead acumulado")
    else:
        axs[0].text(0.5, 0.5, "Sin niveles resueltos",
                    ha="center", va="center", fontsize=11)
        axs[0].axis("off")

    # Panel derecho: desglose por función × nivel (stacked bar)
    top_syms = [s for s in symbols if s["levels"]][:5]
    if top_syms:
        all_lvls_seen = sorted(
            {lv for s in top_syms for lv in s["levels"]},
            key=lambda x: {"L1 hit": 0, "L2 hit": 1, "L3 hit": 2,
                           "Local RAM": 3, "Remote RAM": 4}.get(x, 99)
        )
        sym_names = [_shorten(s["symbol"], 28) for s in top_syms]
        palette2  = ["#70AD47","#1F9EB7","#2E75B6","#7030A0","#C00000"]
        bottoms   = [0.0] * len(top_syms)

        for ci, lvl in enumerate(all_lvls_seen):
            vals = [s["levels"].get(lvl, 0.0) for s in top_syms]
            axs[1].barh(sym_names, vals, left=bottoms,
                        color=palette2[ci % len(palette2)], label=lvl)
            bottoms = [b + v for b, v in zip(bottoms, vals)]

        axs[1].set_title("Accesos resueltos por función y nivel")
        axs[1].set_xlabel("% overhead")
        axs[1].legend(fontsize=8, loc="lower right")
    else:
        axs[1].text(0.5, 0.5,
                    "Sin accesos con nivel resuelto\n(muestras N/A dominantes)",
                    ha="center", va="center", fontsize=10)
        axs[1].axis("off")

    fig.suptitle("perf mem — jerarquía de accesos a memoria (IBS/AMD)",
                 fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "11_perf_mem.png")

def chart_memory_focus(df: pd.DataFrame) -> str:
    sizes = df["road_length"].astype(int).tolist()
    pos, labels = _cat_axis(sizes)

    fig, axs = plt.subplots(2, 2, figsize=(13, 8))
    charts = [
        ("cache_miss_pct", "Cache misses globales", "#C00000"),
        ("l1_miss_pct",    "L1-D miss rate",        "#ED7D31"),
        ("llc_miss_pct",   "LLC miss rate",         "#70AD47"),
        ("dtlb_miss_pct",  "dTLB miss rate",        "#2F5597"),
    ]

    for ax, (col, title, color) in zip(axs.flat, charts):
        if col in df.columns and df[col].notna().any():
            ax.plot(pos, df[col], marker="o", lw=2.2, color=color)
            annotate_points(ax, pos, df[col], d=2)
            ax.set_title(title)
            ax.set_ylabel("%")
        _set_cat_xticks(ax, pos, labels)

    fig.suptitle("Jerarquía de memoria — foco principal del análisis",
                 fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "03_memory_focus.png")


def chart_work_volume(df: pd.DataFrame) -> str:
    sizes = df["road_length"].astype(int).tolist()
    pos, labels = _cat_axis(sizes)

    fig, axs = plt.subplots(2, 2, figsize=(13, 8))

    if "instructions" in df.columns and df["instructions"].notna().any():
        axs[0, 0].plot(pos, df["instructions"] / 1e9, marker="o", lw=2.3, color="#2E75B6")
        axs[0, 0].set_title("Instructions")
        axs[0, 0].set_ylabel("Billions")

    if "cycles" in df.columns and df["cycles"].notna().any():
        axs[0, 1].plot(pos, df["cycles"] / 1e9, marker="o", lw=2.3, color="#7030A0")
        axs[0, 1].set_title("Cycles")
        axs[0, 1].set_ylabel("Billions")

    if "cache_refs" in df.columns and df["cache_refs"].notna().any():
        axs[1, 0].plot(pos, df["cache_refs"] / 1e9, marker="o", lw=2.3, color="#C00000")
        axs[1, 0].set_title("Cache references")
        axs[1, 0].set_ylabel("Billions")

    if "ipc" in df.columns and df["ipc"].notna().any():
        axs[1, 1].plot(pos, df["ipc"], marker="o", lw=2.3, color="#1F9EB7")
        annotate_points(axs[1, 1], pos, df["ipc"], d=3)
        axs[1, 1].set_title("IPC")
        axs[1, 1].set_ylabel("IPC")

    for ax in axs.flat:
        _set_cat_xticks(ax, pos, labels)

    fig.suptitle("Volumen de trabajo del kernel",
                 fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "04_work_volume.png")


def _barh(ax, rows, label_key, value_key, title, color, xlabel="%"):
    if not rows:
        ax.text(0.5, 0.5, "Sin datos", ha="center", va="center", fontsize=11)
        ax.axis("off")
        return
    rows   = rows[:6]
    labels = [_shorten(str(r[label_key]), 34) for r in rows][::-1]
    values = [float(r[value_key]) for r in rows][::-1]
    bars   = ax.barh(labels, values, color=color)
    mx     = max(values) if values else 1
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + mx * 0.02, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}", va="center", fontsize=8)
    ax.set_title(title)
    ax.set_xlabel(xlabel)


def chart_hotspots(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> str:
    largest_n  = int(df["road_length"].max())
    gprof_rows = raw.get(largest_n, {}).get("gprof", {}).get("top", [])
    perf_rows  = raw.get(largest_n, {}).get("perf_record", {}).get("top", [])
    cache_rows = raw.get(largest_n, {}).get("cachegrind", {}).get("top_functions", [])

    fig, axs = plt.subplots(1, 3, figsize=(17, 5.8))
    _barh(axs[0], gprof_rows,  "name",   "pct_time",     f"gprof (N={largest_n:,})",   "#C00000")
    _barh(axs[1], perf_rows,   "symbol", "overhead_pct", "perf record",                 "#2E75B6")
    _barh(axs[2], cache_rows,  "name",   "pct",          "cachegrind instruction share","#70AD47")

    fig.suptitle("Hotspots principales del caso más grande",
                 fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "05_hotspots.png")


def chart_hotspot_inner_gprof(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> str:
    largest_n = int(df["road_length"].max())
    gc = raw.get(largest_n, {}).get("gprof_call", {})
    funcs = gc.get("functions", [])

    fig, axs = plt.subplots(1, 2, figsize=(15, 5.6))

    if funcs:
        top = funcs[0]
        children = sorted(top.get("children", []),
                          key=lambda x: x["self_s"] + x["children_s"],
                          reverse=True)[:8]
        parents = sorted(top.get("parents", []),
                         key=lambda x: x["self_s"] + x["children_s"],
                         reverse=True)[:8]

        child_rows = [
            {"name": c["name"], "time": c["self_s"] + c["children_s"]}
            for c in children if c["name"] != "<spontaneous>"
        ]
        parent_rows = [
            {"name": p["name"], "time": p["self_s"] + p["children_s"]}
            for p in parents if p["name"] != "<spontaneous>"
        ]

        _barh(axs[0], child_rows, "name", "time",
              f"Hijas de {_shorten(top['name'], 24)}", "#2E75B6", xlabel="s")
        _barh(axs[1], parent_rows, "name", "time",
              f"Padres de {_shorten(top['name'], 24)}", "#ED7D31", xlabel="s")
    else:
        for ax in axs:
            ax.text(0.5, 0.5, "Sin call graph usable de gprof",
                    ha="center", va="center", fontsize=11)
            ax.axis("off")

    fig.suptitle("Relaciones internas del hotspot según gprof",
                 fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "06_hotspot_inner_gprof.png")


def chart_hotspot_inner_perf(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> str:
    largest_n = int(df["road_length"].max())
    perf_top = raw.get(largest_n, {}).get("perf_record", {}).get("top", [])
    perf_hier = raw.get(largest_n, {}).get("perf_hier", {}).get("groups", [])

    fig, axs = plt.subplots(1, 2, figsize=(15, 5.6))

    _barh(axs[0], perf_top[:8], "symbol", "overhead_pct",
          f"Símbolos top perf (N={largest_n:,})", "#7030A0")

    if perf_hier:
        _barh(axs[1], perf_hier[:8], "symbol", "overhead_pct",
              "Agrupación consolidada de perf", "#1F9EB7")
    else:
        axs[1].text(0.5, 0.5, "Sin jerarquía útil de perf",
                    ha="center", va="center", fontsize=11)
        axs[1].axis("off")

    fig.suptitle("Desglose interno del hotspot según perf",
                 fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "07_hotspot_inner_perf.png")


def chart_hot_lines(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> str:
    largest_n = int(df["road_length"].max())
    hot = raw.get(largest_n, {}).get("cachegrind", {}).get("hot_lines", [])[:8]

    fig, ax = plt.subplots(figsize=(12, 5.8))

    if hot:
        labels = [_shorten(x["code"], 68) for x in hot][::-1]
        values = [x["pct"] for x in hot][::-1]
        bars = ax.barh(labels, values, color="#C00000")
        mx = max(values) if values else 1
        for bar, val in zip(bars, values):
            ax.text(bar.get_width() + mx * 0.02, bar.get_y() + bar.get_height()/2,
                    f"{val:.2f}", va="center", fontsize=8)
        ax.set_xlabel("% Ir")
        ax.set_title(f"Líneas más calientes de cachegrind (N={largest_n:,})")
    else:
        ax.text(0.5, 0.5, "Sin líneas anotadas",
                ha="center", va="center", fontsize=12)
        ax.axis("off")

    fig.suptitle("Detalle por línea del kernel",
                 fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "08_hot_lines.png")


def chart_hotspot_across_sizes(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> str:
    sizes = df["road_length"].astype(int).tolist()
    pos, labels = _cat_axis(sizes)

    gprof_vals = []
    perf_vals = []
    names = []

    for n in sizes:
        entry = raw.get(n, {})
        gtop = entry.get("gprof", {}).get("top", [])
        ptop = entry.get("perf_record", {}).get("top", [])
        gprof_vals.append(gtop[0]["pct_time"] if gtop else math.nan)
        perf_vals.append(ptop[0]["overhead_pct"] if ptop else math.nan)
        names.append(_best_hot_name(entry))

    fig, axs = plt.subplots(2, 1, figsize=(12, 8), sharex=True)

    axs[0].plot(pos, gprof_vals, marker="o", lw=2.3, color="#C00000", label="gprof top %")
    axs[0].plot(pos, perf_vals, marker="s", lw=2.3, color="#2E75B6", label="perf top %")
    axs[0].set_ylabel("%")
    axs[0].set_title("Concentración del hotspot principal por tamaño")
    axs[0].legend()

    y = [1] * len(pos)
    axs[1].scatter(pos, y, s=60, color="#7030A0")
    for x, name in zip(pos, names):
        axs[1].annotate(_shorten(name, 26), (x, 1), textcoords="offset points",
                        xytext=(0, 8), ha="center", fontsize=8)
    axs[1].set_ylim(0.8, 1.25)
    axs[1].set_yticks([])
    axs[1].set_title("Nombre del hotspot dominante por tamaño")

    _set_cat_xticks(axs[1], pos, labels)
    _set_cat_xticks(axs[0], pos, labels)

    fig.suptitle("Evolución del hotspot dominante",
                 fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "09_hotspot_across_sizes.png")


def chart_timing_variability(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> str:
    largest_n = int(df["road_length"].max())
    timing = raw.get(largest_n, {}).get("timing", {})
    values = timing.get("values", [])

    fig, axs = plt.subplots(1, 2, figsize=(12, 4.8))
    if values:
        xs = list(range(1, len(values) + 1))
        axs[0].plot(xs, values, marker="o", lw=2.1, color="#2E75B6")
        axs[0].axhline(statistics.mean(values), color="#C00000", lw=1.2, ls="--",
                       label=f"mean={statistics.mean(values):.3f} ms")
        axs[0].legend(fontsize=9)
        axs[0].set_title(f"Corridas de tiempo (N={largest_n:,})")
        axs[0].set_xlabel("Run")
        axs[0].set_ylabel("ms")

        axs[1].boxplot(values, vert=True, patch_artist=True,
                       boxprops=dict(facecolor="#DCE6F1"))
        axs[1].scatter([1] * len(values), values, color="#2E75B6", alpha=0.75)
        axs[1].set_title("Dispersión temporal")
        axs[1].set_ylabel("ms")
        axs[1].set_xticks([1])
        axs[1].set_xticklabels([f"N={largest_n:,}"])
    else:
        for ax in axs:
            ax.text(0.5, 0.5, "Sin timing_ms.txt", ha="center", va="center", fontsize=12)
            ax.axis("off")

    fig.suptitle("Variabilidad de corridas",
                 fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "10_timing_variability.png")


def build_chart_pack(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> list[str]:
    os.makedirs(CHARTS_DIR, exist_ok=True)
    paths = [
        chart_performance(df),
        chart_velocity(df),
        chart_memory_focus(df),
        chart_perf_mem(df, raw),
        chart_work_volume(df),
        chart_hotspots(df, raw),
        chart_hotspot_inner_gprof(df, raw),
        chart_hotspot_inner_perf(df, raw),
        chart_hot_lines(df, raw),
        chart_hotspot_across_sizes(df, raw),
        chart_timing_variability(df, raw),
    ]
    return [p for p in paths if p and os.path.exists(p)]


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------
def _write_overview(wb: Workbook, df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Overview")
    _title_row(ws, "Profiling Report — Traffic Automaton Sequential Implementation", 4)

    for col, w in enumerate([28, 22, 30, 68], 1):
        _set_width(ws, col, w)

    for col, label in enumerate(["Indicador", "Contexto", "Valor", "Interpretación"], 1):
        _hcell(ws.cell(3, col), label, "blue")
    ws.freeze_panes = "A4"

    largest_n = int(df["road_length"].max())
    last      = df[df["road_length"] == largest_n].iloc[0]
    timing    = raw.get(largest_n, {}).get("timing", {})
    gprof_top = raw.get(largest_n, {}).get("gprof", {}).get("top", [])
    perf_top  = raw.get(largest_n, {}).get("perf_record", {}).get("top", [])
    gcall     = raw.get(largest_n, {}).get("gprof_call", {}).get("functions", [])
    massif    = raw.get(largest_n, {}).get("massif", {})

    rows_data = []
    if df["throughput_mcells_s"].notna().any():
        best = df.loc[df["throughput_mcells_s"].idxmax()]
        rows_data.append(("Mejor throughput", f"N={int(best['road_length']):,}",
                          f"{_fn(best['throughput_mcells_s'], 2)} Mcells/s",
                          "Relación N×steps / tiempo de medición."))

    if df["ipc"].notna().any():
        best = df.loc[df["ipc"].idxmax()]
        rows_data.append(("Mejor IPC", f"N={int(best['road_length']):,}",
                          f"{_fn(best['ipc'], 4)} ({_classify_ipc(_sf(best['ipc']))})",
                          "Útil para distinguir límite computacional vs memoria."))

    if timing.get("count", 0):
        rows_data.append(("Variabilidad temporal",
                          f"N={largest_n:,} ({timing['count']} runs)",
                          f"CV={_fp(timing['cv_pct'])} | rango={_fn(timing['min_ms'])}–{_fn(timing['max_ms'])} ms",
                          "CV pequeño sugiere medición estable."))

    if gprof_top:
        rows_data.append(("Hotspot gprof", f"N={largest_n:,}",
                          f"{gprof_top[0]['name']} → {_fp(gprof_top[0]['pct_time'])}",
                          "Función dominante según flat profile."))

    if perf_top:
        rows_data.append(("Hotspot perf", f"N={largest_n:,}",
                          f"{perf_top[0]['symbol']} → {_fp(perf_top[0]['overhead_pct'])}",
                          "Confirmación por muestreo estadístico."))

    if gcall:
        top = gcall[0]
        rows_data.append(("Call graph gprof", f"N={largest_n:,}",
                          f"{top['name']} | hijas={len(top['children'])} | padres={len(top['parents'])}",
                          "Permite analizar el hotspot a nivel interno."))

    if massif and not math.isnan(_sf(massif.get("peak_total_mb"))):
        rows_data.append(("Massif peak", f"N={largest_n:,}",
                          f"{_fn(massif.get('peak_total_mb'), 3)} MB | útil={_fp(massif.get('peak_useful_pct'))}",
                          "Coherencia con el tamaño teórico del working set."))

    if "llc_miss_pct" in df.columns and df["llc_miss_pct"].notna().any():
        row_max = df.loc[df["llc_miss_pct"].idxmax()]
        rows_data.append(("LLC miss rate máximo", f"N={int(row_max['road_length']):,}",
                          _fp(row_max["llc_miss_pct"]),
                          "Indica presión sobre la memoria de último nivel."))

    r = 4
    for label, ctx, val, note in rows_data:
        bg = "alt" if r % 2 == 0 else "white"
        _dcell(ws.cell(r, 1), label, bg=bg, align="left", bold=True)
        _dcell(ws.cell(r, 2), ctx,   bg=bg, align="center")
        _dcell(ws.cell(r, 3), val,   bg=bg, align="center")
        _dcell(ws.cell(r, 4), note,  bg=bg, align="left")
        r += 1


def _write_metrics(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Metrics")
    _title_row(ws, "All Metrics from data_profiling.csv", len(df.columns))
    ws.freeze_panes = "A3"

    for ci, col in enumerate(df.columns, 1):
        _hcell(ws.cell(2, ci), col, "blue")
        _set_width(ws, ci, max(12, min(22, len(col) + 2)))

    for ri, row in enumerate(df.itertuples(index=False), 3):
        bg = "alt" if ri % 2 == 0 else "white"
        for ci, col in enumerate(df.columns, 1):
            val = getattr(row, col)
            if col == "road_length":
                _dcell(ws.cell(ri, ci), _si(val), fmt="#,##0", bg=bg, align="center")
            elif not pd.isna(val):
                fmt = ("0.00" if "pct" in col
                       else "#,##0" if any(k in col for k in ["instructions", "cycles", "refs", "misses"])
                       else "0.000")
                _dcell(ws.cell(ri, ci), float(val), fmt=fmt, bg=bg)
            else:
                _dcell(ws.cell(ri, ci), "—", bg=bg, align="center")


def _write_raw_summary(wb: Workbook, df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Raw Summary")
    headers = [
        "N", "Timing runs", "Mean (ms)", "Std (ms)", "CV %",
        "Mean velocity", "gprof top fn", "gprof %",
        "perf top sym", "perf %", "gprof call top",
        "cachegrind top", "cachegrind %", "massif peak MB"
    ]
    _title_row(ws, "Summaries parsed from raw reports", len(headers))
    ws.freeze_panes = "A3"

    widths = [10, 12, 14, 12, 10, 14, 28, 10, 30, 10, 30, 30, 12, 16]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        _hcell(ws.cell(2, i), h, "blue")
        _set_width(ws, i, w)

    r = 3
    for n in df["road_length"].astype(int).tolist():
        entry      = raw.get(n, {})
        timing     = entry.get("timing", {})
        gprof      = entry.get("gprof", {}).get("top", [])
        perf       = entry.get("perf_record", {}).get("top", [])
        gcall      = entry.get("gprof_call", {}).get("functions", [])
        cache      = entry.get("cachegrind", {})
        cache_top  = cache.get("top_functions", [])
        massif     = entry.get("massif", {})
        bg = "alt" if r % 2 == 0 else "white"

        values = [
            n,
            timing.get("count", 0),
            timing.get("mean_ms", math.nan),
            timing.get("std_ms", math.nan),
            timing.get("cv_pct", math.nan),
            timing.get("mean_vel", math.nan),
            gprof[0]["name"] if gprof else "—",
            gprof[0]["pct_time"] if gprof else math.nan,
            perf[0]["symbol"] if perf else "—",
            perf[0]["overhead_pct"] if perf else math.nan,
            gcall[0]["name"] if gcall else "—",
            cache_top[0]["name"] if cache_top else "—",
            cache_top[0]["pct"] if cache_top else math.nan,
            massif.get("peak_total_mb", math.nan),
        ]
        fmts = ["#,##0", "0", "0.000", "0.000", "0.00", "0.000000",
                None, "0.00", None, "0.00", None, None, "0.00", "0.000"]

        for ci, (val, fmt) in enumerate(zip(values, fmts), 1):
            if isinstance(val, str):
                _dcell(ws.cell(r, ci), val, bg=bg, align="left")
            elif fmt:
                _dcell(ws.cell(r, ci), val, fmt=fmt, bg=bg,
                       align="center" if ci <= 2 else "right")
            else:
                _dcell(ws.cell(r, ci), val, bg=bg)
        r += 1


def _write_gprof_callgraph(wb: Workbook, df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Gprof Call Graph")
    headers = ["N", "Función centro", "Pct time", "Self s", "Children s", "Padres", "Hijas"]
    _title_row(ws, "Parsed gprof call graph", len(headers))
    ws.freeze_panes = "A3"

    widths = [10, 34, 12, 12, 12, 42, 42]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        _hcell(ws.cell(2, i), h, "blue")
        _set_width(ws, i, w)

    r = 3
    for n in df["road_length"].astype(int).tolist():
        funcs = raw.get(n, {}).get("gprof_call", {}).get("functions", [])
        if not funcs:
            bg = "alt" if r % 2 == 0 else "white"
            _dcell(ws.cell(r, 1), n, fmt="#,##0", bg=bg, align="center")
            _dcell(ws.cell(r, 2), "Sin datos", bg=bg, align="left")
            for c in range(3, 8):
                _dcell(ws.cell(r, c), "—", bg=bg, align="center")
            r += 1
            continue

        for f in funcs[:6]:
            bg = "alt" if r % 2 == 0 else "white"
            parents = ", ".join(_shorten(p["name"], 18) for p in f["parents"][:4]) or "—"
            children = ", ".join(_shorten(c["name"], 18) for c in f["children"][:4]) or "—"

            _dcell(ws.cell(r, 1), n, fmt="#,##0", bg=bg, align="center")
            _dcell(ws.cell(r, 2), f["name"], bg=bg, align="left")
            _dcell(ws.cell(r, 3), f["pct_time"], fmt="0.00", bg=bg)
            _dcell(ws.cell(r, 4), f["self_s"], fmt="0.000", bg=bg)
            _dcell(ws.cell(r, 5), f["children_s"], fmt="0.000", bg=bg)
            _dcell(ws.cell(r, 6), parents, bg=bg, align="left")
            _dcell(ws.cell(r, 7), children, bg=bg, align="left")
            r += 1


def _write_hot_lines(wb: Workbook, df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Hot Lines")
    _title_row(ws, "Top annotated source lines from cachegrind", 4)
    ws.freeze_panes = "A3"

    for i, (h, w) in enumerate(zip(["N", "Ir", "Share %", "Source line"],
                                   [10, 18, 12, 95]), 1):
        _hcell(ws.cell(2, i), h, "blue")
        _set_width(ws, i, w)

    r = 3
    for n in df["road_length"].astype(int).tolist():
        hot = raw.get(n, {}).get("cachegrind", {}).get("hot_lines", [])
        if not hot:
            bg = "alt" if r % 2 == 0 else "white"
            _dcell(ws.cell(r, 1), n, fmt="#,##0", bg=bg, align="center")
            for ci in (2, 3):
                _dcell(ws.cell(r, ci), "—", bg=bg, align="center")
            _dcell(ws.cell(r, 4), "Sin líneas anotadas", bg=bg, align="left")
            r += 1
            continue

        for item in hot:
            bg = "alt" if r % 2 == 0 else "white"
            _dcell(ws.cell(r, 1), n, fmt="#,##0", bg=bg, align="center")
            _dcell(ws.cell(r, 2), item["ir"], fmt="#,##0", bg=bg)
            _dcell(ws.cell(r, 3), item["pct"], fmt="0.00", bg=bg)
            _dcell(ws.cell(r, 4), item["code"], bg=bg, align="left")
            r += 1


def _write_analysis_notes(wb: Workbook, df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Analysis Notes")
    _title_row(ws, "Auto-generated interpretation notes", 3)

    for i, w in enumerate([28, 22, 92], 1):
        _set_width(ws, i, w)
    for i, h in enumerate(["Tema", "Contexto", "Comentario"], 1):
        _hcell(ws.cell(2, i), h, "blue")

    largest_n = int(df["road_length"].max())
    last      = df[df["road_length"] == largest_n].iloc[0]
    timing    = raw.get(largest_n, {}).get("timing", {})
    gprof_top = raw.get(largest_n, {}).get("gprof", {}).get("top", [])
    perf_top  = raw.get(largest_n, {}).get("perf_record", {}).get("top", [])
    gcall     = raw.get(largest_n, {}).get("gprof_call", {}).get("functions", [])
    cache_top = raw.get(largest_n, {}).get("cachegrind", {}).get("top_functions", [])

    notes = [
        ("Modelo de acceso", "Autómata 1D",
         "El kernel accede a arrays de forma secuencial, por lo que el análisis de memoria "
         "debe priorizar L1/LLC/dTLB antes que branch prediction."),
        ("Working set", "Función de N",
         "El working set del paso temporal es aproximadamente 2×N×4 bytes; al crecer N, "
         "el comportamiento pasa de caché privada a caché compartida o DRAM."),
        ("IPC", f"N={largest_n:,}",
         f"IPC={_fn(last.get('ipc'), 4)} → {_classify_ipc(_sf(last.get('ipc')))}. "
         "Un IPC bajo en este kernel suele reflejar espera por memoria."),
    ]

    if timing.get("count", 0):
        notes.append(("Estabilidad", f"N={largest_n:,}",
                      f"Media={_fn(timing.get('mean_ms'), 3)} ms, std={_fn(timing.get('std_ms'), 3)} ms, "
                      f"CV={_fp(timing.get('cv_pct'))}."))
    if gprof_top:
        notes.append(("Hotspot gprof", f"N={largest_n:,}",
                      f"{gprof_top[0]['name']} concentra {_fp(gprof_top[0]['pct_time'])} del tiempo muestral."))
    if perf_top:
        notes.append(("Hotspot perf", f"N={largest_n:,}",
                      f"{perf_top[0]['symbol']} concentra {_fp(perf_top[0]['overhead_pct'])} del overhead."))
    if gcall:
        notes.append(("Call graph", f"N={largest_n:,}",
                      f"La función central {_shorten(gcall[0]['name'], 28)} tiene "
                      f"{len(gcall[0]['parents'])} padres y {len(gcall[0]['children'])} hijas visibles."))
    if cache_top:
        notes.append(("Cachegrind", f"N={largest_n:,}",
                      f"{cache_top[0]['name']} domina por conteo de instrucciones con {_fp(cache_top[0]['pct'])}."))

    r = 3
    for topic, ctx, note in notes:
        bg = "alt" if r % 2 == 0 else "white"
        _dcell(ws.cell(r, 1), topic, bg=bg, align="left", bold=True)
        _dcell(ws.cell(r, 2), ctx,   bg=bg, align="center")
        _dcell(ws.cell(r, 3), note,  bg=bg, align="left")
        r += 1


def _write_charts_sheet(wb: Workbook, chart_paths: list[str]) -> None:
    ws = wb.create_sheet("Charts")
    _title_row(ws, "Visual report", 20)

    anchors = ["A3", "L3", "A28", "L28", "A53", "L53", "A78", "L78", "A103", "L103"]
    for anchor, path in zip(anchors, chart_paths):
        if not os.path.exists(path):
            continue
        img = XLImage(path)
        img.width = 760
        img.height = 430
        ws.add_image(img, anchor)


def build_workbook(df: pd.DataFrame, raw: dict[int, dict[str, Any]], chart_paths: list[str]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _write_overview(wb, df, raw)
    _write_metrics(wb, df)
    _write_raw_summary(wb, df, raw)
    _write_gprof_callgraph(wb, df, raw)
    _write_hot_lines(wb, df, raw)
    _write_analysis_notes(wb, df, raw)
    _write_charts_sheet(wb, chart_paths)

    wb.save(XLSX_PATH)


# ---------------------------------------------------------------------------
# LaTeX
# ---------------------------------------------------------------------------
def build_latex_table(df: pd.DataFrame) -> None:
    lines = [
        r"\begin{table}[h]",
        r"\centering",
        r"\caption{Caracterización de la implementación secuencial del autómata de tráfico ($\rho=0.50$)}",
        r"\label{tab:profiling_seq}",
        r"\begin{tabular}{rrrrrrrr}",
        r"\hline",
        r"$N$ & $t$ (ms) & $\sigma_t$ (ms) & Mcells/s & IPC & L1 miss & LLC miss & Heap MB \\",
        r"\hline",
    ]
    for _, row in df.iterrows():
        heapval = _sf(row.get("peak_heap_mb"))
        heapstr = f"{heapval:.3f}" if not math.isnan(heapval) else "---"
        lines.append(
            f"{int(row['road_length'])} & "
            f"{_sf(row.get('time_mean_ms')):.3f} & "
            f"{_sf(row.get('time_std_ms')):.3f} & "
            f"{_sf(row.get('throughput_mcells_s')):.3f} & "
            f"{_sf(row.get('ipc')):.4f} & "
            f"{_sf(row.get('l1_miss_pct')):.2f} & "
            f"{_sf(row.get('llc_miss_pct')):.2f} & "
            f"{heapstr} \\\\"
        )
    lines += [r"\hline", r"\end{tabular}", r"\end{table}"]

    with open(LATEX_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------
CSS = """
:root{
  --bg:#f4f7fb; --surface:#fff; --text:#1f2a36; --muted:#5d6b79;
  --border:#d8e1ea; --primary:#1f3557; --accent:#2e75b6;
  --shadow:0 10px 30px rgba(31,53,87,.08);
}
*{box-sizing:border-box}
body{
  margin:0; font-family:Inter,Segoe UI,Arial,sans-serif; color:var(--text);
  background:linear-gradient(180deg,#f8fbff 0,var(--bg) 100%); line-height:1.55;
}
.container{width:min(1380px,94vw);margin:0 auto;padding:32px 0 48px}
.hero{
  background:linear-gradient(135deg,#1f3557,#2e75b6);color:#fff;border-radius:22px;
  padding:36px;box-shadow:var(--shadow);margin-bottom:26px;
}
.hero h1{margin:0 0 8px;font-size:2.1rem}
.hero p{margin:0;color:rgba(255,255,255,.88);max-width:82ch}
.badge{
  background:rgba(255,255,255,.14);border:1px solid rgba(255,255,255,.18);
  padding:6px 10px;border-radius:999px;font-size:.9rem;display:inline-block;margin:4px 4px 0 0;
}
.grid{display:grid;gap:18px}
.kpis{grid-template-columns:repeat(auto-fit,minmax(220px,1fr));margin-bottom:24px}
.two{grid-template-columns:repeat(auto-fit,minmax(420px,1fr));margin-bottom:24px}
.charts{grid-template-columns:repeat(auto-fit,minmax(480px,1fr))}
.card{
  background:var(--surface);border:1px solid var(--border);border-radius:18px;
  box-shadow:var(--shadow);padding:20px;
}
.kpi-label{font-size:.8rem;text-transform:uppercase;letter-spacing:.08em;color:var(--muted);margin-bottom:8px}
.kpi-value{font-size:1.4rem;font-weight:700;color:var(--primary);margin-bottom:4px}
.kpi-note{color:var(--muted);font-size:.94rem}
h2,h3{margin-top:0;color:var(--primary)}
table{width:100%;border-collapse:collapse;font-size:.94rem}
th,td{padding:10px 12px;border-bottom:1px solid var(--border);text-align:left;vertical-align:top}
th{background:#f9fbfd;color:var(--primary)}
.chart-card img{width:100%;height:auto;border-radius:12px;display:block}
.section{margin-top:26px}
"""


def build_html_report(df: pd.DataFrame, raw: dict[int, dict[str, Any]], chart_paths: list[str]) -> None:
    largest_n = int(df["road_length"].max())
    last      = df[df["road_length"] == largest_n].iloc[0]
    timing    = raw.get(largest_n, {}).get("timing", {})
    gprof_top = raw.get(largest_n, {}).get("gprof", {}).get("top", [])
    perf_top  = raw.get(largest_n, {}).get("perf_record", {}).get("top", [])
    gcall     = raw.get(largest_n, {}).get("gprof_call", {}).get("functions", [])

    def kpi(label, value, note):
        return (
            f"<div class='card'>"
            f"<div class='kpi-label'>{html.escape(label)}</div>"
            f"<div class='kpi-value'>{html.escape(value)}</div>"
            f"<div class='kpi-note'>{html.escape(note)}</div>"
            f"</div>"
        )

    def htable(title, rows, nk, vk):
        if not rows:
            return f"<h3>{html.escape(title)}</h3><p>Sin datos.</p>"
        body = "".join(
            f"<tr><td>{html.escape(str(r.get(nk,'')))}</td><td>{_fp(r.get(vk, math.nan))}</td></tr>"
            for r in rows[:8]
        )
        return (
            f"<h3>{html.escape(title)}</h3>"
            f"<table><tr><th>Función / símbolo</th><th>Share</th></tr>{body}</table>"
        )

    def img_tag(path):
        if not os.path.exists(path):
            return ""
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        return f"<div class='card chart-card'><img src='data:image/png;base64,{b64}' alt='chart'></div>"

    cards_html = "".join([
        kpi("Largest N tested", f"N={largest_n:,}",
            f"time={_fn(last.get('time_mean_ms'))} ms | IPC={_fn(last.get('ipc'), 4)}"),
        kpi("Peak throughput",
            f"{_fn(df['throughput_mcells_s'].max() if 'throughput_mcells_s' in df.columns else math.nan, 2)} Mcells/s",
            "N×steps / wall time"),
        kpi("Hotspot gprof",
            gprof_top[0]["name"] if gprof_top else "—",
            _fp(gprof_top[0]["pct_time"]) if gprof_top else "—"),
        kpi("Hotspot perf",
            perf_top[0]["symbol"] if perf_top else "—",
            _fp(perf_top[0]["overhead_pct"]) if perf_top else "—"),
        kpi("Call graph center",
            gcall[0]["name"] if gcall else "—",
            f"padres={len(gcall[0]['parents'])}, hijas={len(gcall[0]['children'])}" if gcall else "—"),
        kpi("Timing stability",
            f"CV={_fp(timing.get('cv_pct'))}",
            f"runs={timing.get('count',0)}"),
    ])

    rows_html = "".join(
        "<tr>" + "".join(
            f"<td>{int(r[c])}</td>" if c == "road_length"
            else f"<td>{_fp(r[c])}</td>" if "pct" in c
            else f"<td>{_fn(r[c],3)}</td>"
            for c in ["road_length", "time_mean_ms", "time_std_ms",
                      "throughput_mcells_s", "ipc", "l1_miss_pct",
                      "llc_miss_pct", "peak_heap_mb"]
        ) + "</tr>"
        for _, r in df.iterrows()
    )

    charts_html = "".join(img_tag(p) for p in chart_paths)

    html_doc = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Profiling Report — Traffic Automaton</title>
<style>{CSS}</style>
</head>
<body>
<div class="container">
  <div class="hero">
    <h1>Profiling Report — Traffic Automaton Sequential</h1>
    <p>Tamaños analizados: {html.escape(', '.join(f'N={int(n):,}' for n in df['road_length'].tolist()))}. Densidad objetivo 0.50.</p>
    <span class="badge">gprof</span>
    <span class="badge">perf stat</span>
    <span class="badge">perf record</span>
    <span class="badge">cachegrind</span>
    <span class="badge">massif</span>
  </div>

  <div class="grid kpis">{cards_html}</div>

  <div class="section card">
    <h2>Métricas por tamaño</h2>
    <table>
      <tr>
        <th>N</th><th>t ms</th><th>σ ms</th><th>Mcells/s</th>
        <th>IPC</th><th>L1 miss %</th><th>LLC miss %</th><th>Heap MB</th>
      </tr>
      {rows_html}
    </table>
  </div>

  <div class="section">
    <h2>Hotspots</h2>
    <div class="grid two">
      <div class="card">{htable("gprof top funciones", gprof_top, "name", "pct_time")}</div>
      <div class="card">{htable("perf top símbolos", perf_top, "symbol", "overhead_pct")}</div>
    </div>
  </div>

  <div class="section">
    <h2>Gráficas</h2>
    <div class="grid charts">{charts_html}</div>
  </div>
</div>
</body>
</html>
"""
    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html_doc)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    print(f"[+] Reading CSV: {CSV_PATH}")
    df = normalize_dataframe(CSV_PATH)
    print(f"[+] {len(df)} rows loaded, N={df['road_length'].tolist()}")

    print("[+] Collecting raw reports")
    raw = collect_raw(df)

    print("[+] Building charts")
    chart_paths = build_chart_pack(df, raw)
    print(f"[+] {len(chart_paths)} charts saved to {CHARTS_DIR}")

    print("[+] Building Excel workbook")
    build_workbook(df, raw, chart_paths)
    print(f"[+] Saved: {XLSX_PATH}")

    print("[+] Building LaTeX table")
    build_latex_table(df)
    print(f"[+] Saved: {LATEX_PATH}")

    print("[+] Building HTML report")
    build_html_report(df, raw, chart_paths)
    print(f"[+] Saved: {HTML_PATH}")

    print("[+] Done.")


if __name__ == "__main__":
    main()