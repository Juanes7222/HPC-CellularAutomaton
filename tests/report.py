"""
report.py  --  Traffic Automaton benchmark report generator  (v5)

Structure of each sheet
-----------------------
  Tabla 1  — Mediciones individuales:
              rows = repetitions, cols = N sizes
              + summary rows: Promedio / Desv. Est. / CV (%)
              One block per implementation, stacked vertically.

  Tabla 2  — Promedio y Speedup:
              One row per implementation.
              Cols: N sizes (avg ms) | N sizes (speedup vs best_seq) | SpUp Prom.

  Two charts per stage (time + speedup), each saved as a separate PNG.

Stages
------
  0. Secuencial       seq | seq_opt | seq_mem | seq_mem_opt
  1. OMP base         best_seq vs omp_p2..p12
  2. OMP compilador   best_seq vs omp_opt_p2..p12
  3. OMP comp+mem     best_seq vs omp_mem_opt_p2..p12
  4. Final            best of each strategy

  5. Densidad         time + velocity vs density  (separate sheet)

CSV columns:
  impl, threads, road_length, density, repetition, wall_time_ms, avg_velocity
"""

from __future__ import annotations
import math, os, sys
from dataclasses import dataclass

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
RESULTS_DIR = sys.argv[1].rstrip("/") if len(sys.argv) > 1 else "tests/machine1/results_traffic"
OUTPUT_PATH = os.path.join(RESULTS_DIR, "reporte_traffic.xlsx")
CHARTS_DIR  = os.path.join(RESULTS_DIR, "charts")
TABLES_DIR  = os.path.join(RESULTS_DIR, "tables")
CSV_FILE    = "data.csv"

MEASURE_STEPS   = 1000
SCALING_DENSITY = 0.50
N_VALUES        = [8_000, 64_000, 500_000, 2_000_000, 5_000_000, 10_000_000]
DENSITY_EXPERIMENT_N       = 20_000_000
DENSITY_EXPERIMENT_THREADS = [0, 4, 8, 12]
DENSITY_VALUES             = [0.10, 0.30, 0.50, 0.70, 0.90]

SEQ         = "Serial Base"
SEQ_OPT     = "Serial + Compiler Opt"
SEQ_MEM     = "Serial + Memory Opt"
SEQ_MEM_OPT = "Serial + Compiler & Memory Opt"
OMP         = "OpenMP Base"
OMP_OPT     = "OpenMP + Compiler Opt"
OMP_MEM     = "OpenMP + Memory Opt"
OMP_MEM_OPT = "OpenMP + Compiler & Memory Opt"

IMPL_NAME_MAP = {
    "traffic_seq":         SEQ,
    "traffic_seq_opt":     SEQ_OPT,
    "traffic_seq_mem":     SEQ_MEM,
    "traffic_seq_mem_opt": SEQ_MEM_OPT,
    "traffic_omp":         OMP,
    "traffic_omp_opt":     OMP_OPT,
    "traffic_omp_mem":     OMP_MEM,
    "traffic_omp_mem_opt": OMP_MEM_OPT,
}

PALETTE = {
    SEQ:         "#2E75B6",
    SEQ_OPT:     "#ED7D31",
    SEQ_MEM:     "#70AD47",
    SEQ_MEM_OPT: "#C00000",
    OMP:         "#7030A0",
    OMP_OPT:     "#00B0F0",
    OMP_MEM:     "#FF66CC",
    OMP_MEM_OPT: "#375623",
}
THREAD_PALETTE = {2: "#BBBBBB", 4: "#70AD47", 6: "#4472C4", 8: "#ED7D31", 12: "#000A94"}

CHART_STYLE = {
    "figure.facecolor": "white", "axes.facecolor": "white",
    "axes.grid": True, "grid.color": "#E7EDF3", "grid.linestyle": "--",
    "grid.alpha": 0.75, "axes.spines.top": False, "axes.spines.right": False,
    "axes.titleweight": "bold", "axes.labelsize": 10, "axes.titlesize": 11,
    "legend.frameon": False, "font.size": 10,
}
FONT_NAME = "Calibri"

C = {
    "dark":       "1F3557",
    "mid":        "2E75B6",
    "light":      "DCE6F1",
    "alt":        "F4F8FC",
    "green_bg":   "E2F0D9",
    "green_fg":   "2F6B1E",
    "summary_bg": "EBF3FB",
    "summary_fg": "1F4E79",
    "impl_hdr":   "1F4E79",
    "sep":        "D9E1F2",
}

# ---------------------------------------------------------------------------
# Data types
# ---------------------------------------------------------------------------
# Impl -> {road_length: [(rep, wall_time_ms, avg_velocity), ...]}
AllReps = dict

@dataclass(frozen=True)
class Impl:
    name: str
    parallelism: int   # 0 = serial

    @property
    def label(self) -> str:
        if self.parallelism == 0:
            return self.name
        return f"{self.name}  (p={self.parallelism})"

    @property
    def short_label(self) -> str:
        return self.name if self.parallelism == 0 else f"{self.name}_p{self.parallelism}"

    @property
    def color(self) -> str:
        return PALETTE.get(self.name, "#888888")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _size_label(n):
    if n >= 1_000_000 and n % 1_000_000 == 0: return f"{n//1_000_000}M"
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000 and n % 1_000 == 0: return f"{n//1_000}K"
    return str(n)

def _mean(vals):
    clean = [float(v) for v in vals if v is not None and not math.isnan(float(v))]
    return sum(clean)/len(clean) if clean else math.nan

def _std(vals):
    clean = [float(v) for v in vals if v is not None and not math.isnan(float(v))]
    if len(clean) < 2: return 0.0
    m = _mean(clean)
    return math.sqrt(sum((v-m)**2 for v in clean)/len(clean))

def _cv(vals):
    m = _mean(vals)
    return 100*_std(vals)/m if m and not math.isnan(m) else math.nan

# ---------------------------------------------------------------------------
# CSV loading
# ---------------------------------------------------------------------------
_CSV_CACHE = None

def load_csv():
    global _CSV_CACHE
    if _CSV_CACHE is not None: return _CSV_CACHE
    path = os.path.join(RESULTS_DIR, CSV_FILE)
    if not os.path.exists(path):
        print(f" [--] {path} not found."); return None
    df = pd.read_csv(path)
    df.columns = [c.strip().lower() for c in df.columns]
    if "threads" in df.columns and "parallelism" not in df.columns:
        df.rename(columns={"threads": "parallelism"}, inplace=True)
    df["impl"]         = df["impl"].str.strip().map(lambda x: IMPL_NAME_MAP.get(x, x))
    df["parallelism"]  = df["parallelism"].astype(int)
    df["road_length"]  = df["road_length"].astype(int)
    df["density"]      = df["density"].astype(float)
    df["repetition"]   = df["repetition"].astype(int)
    df["wall_time_ms"] = df["wall_time_ms"].astype(float)
    df["avg_velocity"] = df["avg_velocity"].astype(float)
    df.loc[df["wall_time_ms"] == 0, "wall_time_ms"] = math.nan
    _CSV_CACHE = df
    print(f" [OK] {path}  ({len(df)} rows)")
    return _CSV_CACHE

def get_scaling_df(df):
    return df[df["road_length"].isin(N_VALUES) & (df["density"] == SCALING_DENSITY)].copy()

def get_density_df(df):
    return df[df["road_length"] == DENSITY_EXPERIMENT_N].copy()

def build_all_reps(df) -> AllReps:
    result = {}
    for (name, par), grp in df.groupby(["impl","parallelism"]):
        impl = Impl(name=str(name), parallelism=int(par))
        result[impl] = {}
        for size, sub in grp.groupby("road_length"):
            result[impl][int(size)] = sorted(
                [(int(r), float(t), float(v))
                 for r, t, v in zip(sub["repetition"], sub["wall_time_ms"], sub["avg_velocity"])
                 if not math.isnan(float(t))],
                key=lambda x: x[0])
    return result

def compute_avgs(all_reps) -> dict:
    return {impl: {s: _mean([t for _,t,_ in pairs]) for s, pairs in sizes.items()}
            for impl, sizes in all_reps.items()}

def find_best_serial(avg_data, sizes):
    serial_names = {SEQ, SEQ_OPT, SEQ_MEM, SEQ_MEM_OPT}
    best_impl, best_sum = None, float("inf")
    for impl, times in avg_data.items():
        if impl.name not in serial_names: continue
        total = sum(times.get(s, float("inf")) for s in sizes)
        if total < best_sum: best_sum = total; best_impl = impl
    return best_impl

def get_omp_variants(avg_data, omp_name):
    return sorted([i for i in avg_data if i.name == omp_name], key=lambda i: i.parallelism)

def best_par(family, avg_data, best_seq, sizes):
    cands = [i for i in avg_data if i.name == family]
    if not cands: return None
    ref_avgs = avg_data.get(best_seq, {})
    best, best_sp = None, 0.0
    for impl in cands:
        sp_vals = [ref_avgs[s]/avg_data[impl][s] for s in sizes
                   if s in avg_data[impl] and s in ref_avgs and avg_data[impl][s] > 0]
        sp = _mean(sp_vals)
        if sp > best_sp: best_sp = sp; best = impl
    return best

# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------
def _thin_border():
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    thick = Side(style="medium", color="1F4E79")
    thin  = Side(style="thin", color="BDD7EE")
    return Border(left=thick, right=thick, top=thin, bottom=thick)

def _hdr(cell, v, bg=C["mid"], fg="FFFFFF", size=10, bold=True, align="center"):
    cell.value     = v
    cell.font      = Font(name=FONT_NAME, bold=bold, size=size, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = _thin_border()

def _dat(cell, v, fmt=None, bg="FFFFFF", fg="000000", bold=False, align="right"):
    cell.value     = v
    cell.font      = Font(name=FONT_NAME, bold=bold, size=10, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _thin_border()
    if fmt: cell.number_format = fmt

def _summary_dat(cell, v, fmt=None, bold=False):
    _dat(cell, v, fmt=fmt, bg=C["summary_bg"], fg=C["summary_fg"], bold=bold)

def _set_w(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def _title_row(ws, text, cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["dark"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26

# ---------------------------------------------------------------------------
# Chart helpers
# ---------------------------------------------------------------------------
def _save_fig(fig, name):
    os.makedirs(CHARTS_DIR, exist_ok=True)
    path = os.path.join(CHARTS_DIR, name)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return path

def _categorical_xticks(ax, sizes):
    positions = list(range(len(sizes)))
    ax.set_xticks(positions)
    ax.set_xticklabels([_size_label(s) for s in sizes], rotation=20, ha="right")
    return positions

def chart_time(impls, avg_data, sizes, title, fname):
    positions = list(range(len(sizes)))
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 5))
        for impl in impls:
            times = avg_data.get(impl, {})
            xs = [positions[i] for i, s in enumerate(sizes) if s in times]
            ys = [times[s] for s in sizes if s in times]
            if xs:
                color = THREAD_PALETTE.get(impl.parallelism, impl.color) if impl.parallelism > 0 else impl.color
                ax.plot(xs, ys, marker="o", lw=2.2, color=color, label=impl.label)
        _categorical_xticks(ax, sizes)
        ax.set_yscale("log")
        ax.set_ylabel("Tiempo promedio (ms)  [log]")
        ax.set_xlabel("N (road length)")
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
    return _save_fig(fig, fname)

def chart_speedup(ref_impl, cmp_impls, avg_data, sizes, title, fname, note=None):
    ref_avgs  = avg_data.get(ref_impl, {})
    positions = list(range(len(sizes)))
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.axhline(1.0, color="#AAAAAA", lw=1.4, ls="--", label=f"ref: {ref_impl.short_label}")
        for impl in cmp_impls:
            if impl not in avg_data: continue
            xs, ys = [], []
            for i, s in enumerate(sizes):
                t = avg_data[impl].get(s); r = ref_avgs.get(s)
                if t and r and t > 0: xs.append(positions[i]); ys.append(r/t)
            if xs:
                color = THREAD_PALETTE.get(impl.parallelism, impl.color) if impl.parallelism > 0 else impl.color
                ax.plot(xs, ys, marker="o", lw=2.2, color=color, label=impl.label)
        _categorical_xticks(ax, sizes)
        ax.set_ylabel("Speedup  T(ref) / T(impl)")
        ax.set_xlabel("N (road length)")
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.legend(fontsize=8, loc="upper left")
        if note:
            ax.annotate(note, xy=(0.01, 0.03), xycoords="axes fraction",
                        fontsize=7, color="#888888", style="italic")
        fig.tight_layout()
    return _save_fig(fig, fname)

def chart_scaling(ref_impl, omp_impls, avg_data, sizes, title, fname):
    ref_avgs   = avg_data.get(ref_impl, {})
    par_counts = sorted({i.parallelism for i in omp_impls if i.parallelism > 0})
    plot_sizes = sizes[:6]
    with plt.rc_context(CHART_STYLE):
        ncols = min(3, len(plot_sizes))
        nrows = math.ceil(len(plot_sizes)/ncols)
        fig, axs = plt.subplots(nrows, ncols, figsize=(5*ncols, 4*nrows), squeeze=False)
        for idx, size in enumerate(plot_sizes):
            ax    = axs[idx//ncols][idx%ncols]
            ref_t = ref_avgs.get(size)
            if not ref_t: ax.set_visible(False); continue
            sp_list, eff_list = [], []
            for p in par_counts:
                impl_here = next((i for i in omp_impls if i.parallelism == p), None)
                t = avg_data.get(impl_here, {}).get(size) if impl_here else None
                if t and t > 0:
                    sp = ref_t/t; sp_list.append(sp); eff_list.append(sp/p*100)
                else:
                    sp_list.append(float("nan")); eff_list.append(float("nan"))
            ax2 = ax.twinx()
            ax2.spines["right"].set_visible(True)
            ax2.spines["right"].set_color("#70AD47")
            ax.plot(par_counts, sp_list,  "o-", color="#2E75B6", lw=2, label="Speedup")
            ax2.plot(par_counts, eff_list, "s--", color="#70AD47", lw=1.5, label="Efic. (%)")
            ax.plot(par_counts, par_counts, "k--", lw=1, alpha=0.3, label="Ideal")
            ax.set_title(f"N={_size_label(size)}", fontsize=10)
            ax.set_xlabel("p (hilos)"); ax.set_ylabel("Speedup")
            ax2.set_ylabel("Eficiencia (%)", color="#70AD47")
            ax.set_xticks(par_counts)
            h1, l1 = ax.get_legend_handles_labels()
            h2, l2 = ax2.get_legend_handles_labels()
            ax.legend(h1+h2, l1+l2, fontsize=7, loc="upper left")
        for idx in range(len(plot_sizes), nrows*ncols):
            axs[idx//ncols][idx%ncols].set_visible(False)
        fig.suptitle(title, fontsize=12, fontweight="bold")
        fig.tight_layout()
    return _save_fig(fig, fname)

def chart_density_time(density_df, title, fname):
    density_values = sorted(density_df["density"].unique())
    all_impls = [(SEQ, 0), (SEQ_OPT, 0), (SEQ_MEM_OPT, 0)] + \
                [(OMP_MEM_OPT, t) for t in DENSITY_EXPERIMENT_THREADS if t > 0]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 5))
        for name, threads in all_impls:
            sub = density_df[(density_df["impl"] == name) & (density_df["parallelism"] == threads)]
            if sub.empty: continue
            lbl   = name if threads == 0 else f"{name} (p={threads})"
            color = THREAD_PALETTE.get(threads, PALETTE.get(name, "#888888"))
            agg   = sub.groupby("density")["wall_time_ms"].mean().reindex(density_values)
            ax.plot(density_values, agg, marker="o", lw=2, color=color, label=lbl)
        ax.set_xlabel("Densidad ρ"); ax.set_ylabel("Tiempo medio (ms)")
        ax.set_xticks(density_values); ax.legend(fontsize=7)
        ax.set_title(title, fontsize=11); fig.tight_layout()
    return _save_fig(fig, fname)

def chart_density_velocity(density_df, title, fname):
    density_values = sorted(density_df["density"].unique())
    all_impls = [(SEQ, 0), (SEQ_OPT, 0), (SEQ_MEM_OPT, 0)] + \
                [(OMP_MEM_OPT, t) for t in DENSITY_EXPERIMENT_THREADS if t > 0]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 5))
        for name, threads in all_impls:
            sub = density_df[(density_df["impl"] == name) & (density_df["parallelism"] == threads)]
            if sub.empty: continue
            lbl   = name if threads == 0 else f"{name} (p={threads})"
            color = THREAD_PALETTE.get(threads, PALETTE.get(name, "#888888"))
            agg   = sub.groupby("density")["avg_velocity"].mean().reindex(density_values)
            ax.plot(density_values, agg, marker="o", lw=2, color=color, label=lbl)
        rho_th = np.linspace(0.01, 0.99, 200)
        v_th   = np.where(rho_th < 0.5, 1.0, (1-rho_th)/rho_th)
        ax.plot(rho_th, v_th, "k--", lw=1.2, alpha=0.5, label="v∞ teórica")
        ax.set_xlabel("Densidad ρ"); ax.set_ylabel("Velocidad media v̄")
        ax.set_xticks(density_values); ax.set_ylim(0, 1.05); ax.legend(fontsize=7)
        ax.set_title(title, fontsize=11); fig.tight_layout()
    return _save_fig(fig, fname)

# ---------------------------------------------------------------------------
# Table 1: raw reps with Promedio / Desv. Est. / CV
# ---------------------------------------------------------------------------
def _write_raw_table(ws, impls, all_reps, sizes, n_cols, start_row):
    cur = start_row

    # Section title
    ws.merge_cells(f"A{cur}:{get_column_letter(n_cols)}{cur}")
    c = ws.cell(cur, 1, value="Tabla 1  —  Mediciones individuales (ms)")
    c.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur].height = 20
    cur += 1

    n_reps = max((len(pairs) for impl in impls
                  for pairs in all_reps.get(impl, {}).values()), default=1)

    for impl_idx, impl in enumerate(impls):
        impl_reps = all_reps.get(impl, {})

        # Implementation block header
        ws.merge_cells(f"A{cur}:{get_column_letter(n_cols)}{cur}")
        bh = ws.cell(cur, 1, value=impl.label)
        bh.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        bh.fill      = PatternFill("solid", fgColor=C["impl_hdr"])
        bh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        bh.border    = _thick_border()
        ws.row_dimensions[cur].height = 20
        cur += 1

        # Column headers
        _hdr(ws.cell(cur, 1), "Rep", bg=C["mid"], size=9)
        for ci, s in enumerate(sizes, 2):
            _hdr(ws.cell(cur, ci), f"N={_size_label(s)}", bg=C["mid"], size=9)
        ws.row_dimensions[cur].height = 18
        cur += 1

        # One row per repetition
        for rep in range(1, n_reps + 1):
            bg = C["alt"] if rep % 2 == 0 else "FFFFFF"
            _dat(ws.cell(cur, 1), rep, bg=C["light"], bold=True, align="center")
            for ci, s in enumerate(sizes, 2):
                val = next((t for r,t,_ in impl_reps.get(s, []) if r == rep), None)
                if val is not None:
                    _dat(ws.cell(cur, ci), round(val, 3), fmt="#,##0.000", bg=bg)
                else:
                    _dat(ws.cell(cur, ci), "—", bg=bg, align="center")
            ws.row_dimensions[cur].height = 16
            cur += 1

        # Summary rows
        for s_label, s_idx in [("Promedio", 0), ("Desv. Est.", 1), ("CV (%)", 2)]:
            _hdr(ws.cell(cur, 1), s_label, bg=C["summary_bg"], fg=C["summary_fg"],
                 size=9, align="left", bold=True)
            for ci, s in enumerate(sizes, 2):
                vals = [t for _,t,_ in impl_reps.get(s, [])]
                if vals:
                    mean = _mean(vals); std = _std(vals); cv = _cv(vals)
                    show = [mean, std, cv][s_idx]
                    fmt  = ["#,##0.000", "#,##0.000", "0.00"][s_idx]
                    _summary_dat(ws.cell(cur, ci), round(show, 3 if s_idx < 2 else 2),
                                 fmt=fmt, bold=(s_idx == 0))
                else:
                    _summary_dat(ws.cell(cur, ci), "—")
            ws.row_dimensions[cur].height = 16
            cur += 1

        # Separator between implementations
        if impl_idx < len(impls) - 1:
            for ci in range(1, n_cols + 1):
                ws.cell(cur, ci).fill = PatternFill("solid", fgColor=C["sep"])
            ws.row_dimensions[cur].height = 6
            cur += 1

    return cur

# ---------------------------------------------------------------------------
# Table 2: average time + speedup per impl
# ---------------------------------------------------------------------------
def _write_speedup_table(ws, impls, avg_data, ref_impl, sizes, start_row):
    cur       = start_row
    n_sp_cols = 1 + len(sizes)*2 + 1

    ws.merge_cells(f"A{cur}:{get_column_letter(n_sp_cols)}{cur}")
    c = ws.cell(cur, 1, value=f"Tabla 2  —  Promedio y Speedup  (ref: {ref_impl.short_label})")
    c.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C["dark"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur].height = 20
    cur += 1

    _set_w(ws, 1, 34)
    for si in range(len(sizes)):
        _set_w(ws, 2 + si*2, 13)
        _set_w(ws, 3 + si*2, 11)
    _set_w(ws, n_sp_cols, 11)

    _hdr(ws.cell(cur, 1), "Implementación", bg=C["dark"], size=10)
    for si, s in enumerate(sizes):
        _hdr(ws.cell(cur, 2+si*2), f"N={_size_label(s)}\nAvg (ms)", bg=C["mid"], size=9)
        _hdr(ws.cell(cur, 3+si*2), f"N={_size_label(s)}\nSpeedup",  bg=C["mid"], size=9)
    _hdr(ws.cell(cur, n_sp_cols), "SpUp\nProm.", bg=C["dark"], size=9)
    ws.row_dimensions[cur].height = 28
    cur += 1

    ref_avgs = avg_data.get(ref_impl, {})
    for ri, impl in enumerate(impls):
        bg    = C["alt"] if ri % 2 == 0 else "FFFFFF"
        times = avg_data.get(impl, {})
        _dat(ws.cell(cur, 1), impl.label, bg=C["light"], bold=True, align="left")

        sp_cells = []
        for si, s in enumerate(sizes):
            avg_t   = times.get(s)
            ref_avg = ref_avgs.get(s)
            ac, sc  = 2 + si*2, 3 + si*2

            c_avg = ws.cell(cur, ac)
            c_avg.value         = round(avg_t, 3) if avg_t else "N/A"
            c_avg.number_format = "#,##0.000"
            c_avg.font          = Font(name=FONT_NAME, size=10)
            c_avg.fill          = PatternFill("solid", fgColor=bg)
            c_avg.alignment     = Alignment(horizontal="right", vertical="center")
            c_avg.border        = _thin_border()

            c_sp = ws.cell(cur, sc)
            if impl == ref_impl:
                c_sp.value = 1.0
            elif avg_t and avg_t > 0 and ref_avg:
                c_sp.value = round(ref_avg / avg_t, 4)
                sp_cells.append(get_column_letter(sc) + str(cur))
            else:
                c_sp.value = "N/A"
            c_sp.number_format = "0.0000"
            c_sp.font          = Font(name=FONT_NAME, size=10, color=C["green_fg"])
            c_sp.fill          = PatternFill("solid", fgColor=C["green_bg"])
            c_sp.alignment     = Alignment(horizontal="right", vertical="center")
            c_sp.border        = _thin_border()

        c_av = ws.cell(cur, n_sp_cols)
        if sp_cells:
            c_av.value = f"=IFERROR(AVERAGE({','.join(sp_cells)}),\"N/A\")"
        elif impl == ref_impl:
            c_av.value = 1.0
        else:
            c_av.value = "N/A"
        c_av.number_format = "0.00"
        c_av.font          = Font(name=FONT_NAME, bold=True, size=10, color=C["green_fg"])
        c_av.fill          = PatternFill("solid", fgColor=C["green_bg"])
        c_av.alignment     = Alignment(horizontal="right", vertical="center")
        c_av.border        = _thin_border()
        ws.row_dimensions[cur].height = 17
        cur += 1

    return cur

# ---------------------------------------------------------------------------
# Full sheet writer
# ---------------------------------------------------------------------------
def write_sheet(wb, sheet_name, title, impls, all_reps, avg_data,
                ref_impl, sizes, chart_time_path, chart_sp_path,
                chart_scale_path=None):
    ws    = wb.create_sheet(sheet_name)
    n_raw = 1 + len(sizes)

    _title_row(ws, title, n_raw)
    _set_w(ws, 1, 8)
    for ci in range(2, n_raw + 1): _set_w(ws, ci, 16)

    cur = _write_raw_table(ws, impls, all_reps, sizes, n_raw, start_row=2)
    cur += 2
    cur = _write_speedup_table(ws, impls, avg_data, ref_impl, sizes, start_row=cur)
    cur += 3

    for anchor, path, w, h in [
        (f"A{cur}",  chart_time_path,  680, 390),
        (f"M{cur}",  chart_sp_path,    680, 390),
        (f"A{cur+24}", chart_scale_path, 1050, 570),
    ]:
        if path and os.path.exists(path):
            img = XLImage(path); img.width = w; img.height = h
            ws.add_image(img, anchor)

# ---------------------------------------------------------------------------
# Density sheet
# ---------------------------------------------------------------------------
def write_density_sheet(wb, density_df, chart_time_path, chart_vel_path):
    ws = wb.create_sheet("5. Densidad")
    density_values = sorted(density_df["density"].unique())
    N_COLS = len(density_values) + 1
    _title_row(ws, f"Experimento de densidad  |  N={_size_label(DENSITY_EXPERIMENT_N)}", N_COLS)
    _set_w(ws, 1, 24)
    for ci in range(2, N_COLS+1): _set_w(ws, ci, 14)

    cur = 2
    for metric, col, fmt in [
        ("Tiempo medio (ms)", "wall_time_ms", "#,##0.000"),
        ("Velocidad media v̄", "avg_velocity", "0.000000"),
    ]:
        ws.merge_cells(f"A{cur}:{get_column_letter(N_COLS)}{cur}")
        bh = ws.cell(cur, 1, value=metric)
        bh.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        bh.fill = PatternFill("solid", fgColor=C["impl_hdr"])
        bh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[cur].height = 20
        cur += 1

        _hdr(ws.cell(cur, 1), "Impl", bg=C["mid"], size=9)
        for ci, d in enumerate(density_values, 2):
            _hdr(ws.cell(cur, ci), f"ρ={d:.2f}", bg=C["mid"], size=9)
        ws.row_dimensions[cur].height = 18
        cur += 1

        for ri, (name, threads) in enumerate([
            (SEQ, 0), (SEQ_OPT, 0), (SEQ_MEM_OPT, 0),
            (OMP_MEM_OPT, 4), (OMP_MEM_OPT, 8), (OMP_MEM_OPT, 12),
        ]):
            sub = density_df[(density_df["impl"]==name) & (density_df["parallelism"]==threads)]
            lbl = name if threads == 0 else f"{name} (p={threads})"
            bg  = C["alt"] if ri % 2 == 0 else "FFFFFF"
            _dat(ws.cell(cur, 1), lbl, bg=C["light"], bold=True, align="left")
            for ci, d in enumerate(density_values, 2):
                val = sub[sub["density"]==d][col].mean() if not sub.empty else math.nan
                _dat(ws.cell(cur, ci),
                     round(val, 4) if not math.isnan(val) else "—",
                     fmt=fmt, bg=bg)
            ws.row_dimensions[cur].height = 16
            cur += 1

        for ci in range(1, N_COLS+1):
            ws.cell(cur, ci).fill = PatternFill("solid", fgColor=C["sep"])
        ws.row_dimensions[cur].height = 8
        cur += 1

    cur += 2
    for anchor, path in [(f"A{cur}", chart_time_path), (f"M{cur}", chart_vel_path)]:
        if path and os.path.exists(path):
            img = XLImage(path); img.width = 680; img.height = 400
            ws.add_image(img, anchor)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    os.makedirs(CHARTS_DIR, exist_ok=True)

    df = load_csv()
    if df is None: print("No data found."); sys.exit(1)

    scaling_df = get_scaling_df(df)
    density_df = get_density_df(df)
    all_reps   = build_all_reps(scaling_df)
    avg_data   = compute_avgs(all_reps)
    sizes      = sorted({s for szs in all_reps.values() for s in szs})

    best_seq = find_best_serial(avg_data, sizes) or Impl(SEQ, 0)
    omp_impls         = get_omp_variants(avg_data, OMP)
    omp_opt_impls     = get_omp_variants(avg_data, OMP_OPT)
    omp_mem_opt_impls = get_omp_variants(avg_data, OMP_MEM_OPT)
    best_omp         = best_par(OMP,         avg_data, best_seq, sizes)
    best_omp_opt     = best_par(OMP_OPT,     avg_data, best_seq, sizes)
    best_omp_mem_opt = best_par(OMP_MEM_OPT, avg_data, best_seq, sizes)

    print(f"\n[INFO] Best serial : {best_seq.short_label}")
    for lbl, impl in [("omp", best_omp), ("omp_opt", best_omp_opt), ("omp_mem_opt", best_omp_mem_opt)]:
        print(f"[INFO] Best {lbl:12s}: {impl.short_label if impl else 'N/A'}")

    print("\nGenerating charts...")

    serial_impls = [Impl(n, 0) for n in [SEQ, SEQ_OPT, SEQ_MEM, SEQ_MEM_OPT] if Impl(n,0) in avg_data]

    ct0  = chart_time(serial_impls, avg_data, sizes, "Secuenciales — tiempo", "fig_seq_time.png")
    cs0  = chart_speedup(Impl(SEQ,0), [i for i in serial_impls if i != Impl(SEQ,0)], avg_data, sizes,
                         "Secuenciales — speedup vs Serial Base", "fig_seq_speedup.png")

    ct1  = chart_time([best_seq]+omp_impls, avg_data, sizes, "OpenMP Base — tiempo", "fig_omp_base_time.png")
    cs1  = chart_speedup(best_seq, omp_impls, avg_data, sizes, "OpenMP Base — speedup", "fig_omp_base_speedup.png",
                         note="p=2 puede ser < 1 para N pequeño")
    csc1 = chart_scaling(best_seq, omp_impls, avg_data, sizes, "OpenMP Base — escalabilidad", "fig_omp_base_scaling.png")

    ct2  = chart_time([best_seq]+omp_opt_impls, avg_data, sizes, "OpenMP + Compiler Opt — tiempo", "fig_omp_comp_time.png")
    cs2  = chart_speedup(best_seq, omp_opt_impls, avg_data, sizes, "OpenMP + Compiler Opt — speedup", "fig_omp_comp_speedup.png")
    csc2 = chart_scaling(best_seq, omp_opt_impls, avg_data, sizes, "OpenMP + Compiler Opt — escalabilidad", "fig_omp_comp_scaling.png")

    ct3  = chart_time([best_seq]+omp_mem_opt_impls, avg_data, sizes, "OpenMP + Compiler & Memory Opt — tiempo", "fig_omp_comp_mem_time.png")
    cs3  = chart_speedup(best_seq, omp_mem_opt_impls, avg_data, sizes, "OpenMP + Compiler & Memory Opt — speedup", "fig_omp_comp_mem_speedup.png")
    csc3 = chart_scaling(best_seq, omp_mem_opt_impls, avg_data, sizes, "OpenMP + Compiler & Memory Opt — escalabilidad", "fig_omp_comp_mem_scaling.png")

    final_impls = [i for i in [best_seq, best_omp, best_omp_opt, best_omp_mem_opt] if i]
    ct4  = chart_time(final_impls, avg_data, sizes, "Comparación final — tiempo", "fig_final_time.png")
    cs4  = chart_speedup(best_seq, [i for i in final_impls if i != best_seq], avg_data, sizes,
                         "Comparación final — speedup", "fig_final_speedup.png")

    c_dt = chart_density_time(density_df,     f"Densidad — tiempo      N={_size_label(DENSITY_EXPERIMENT_N)}", "fig_density_time.png") if not density_df.empty else None
    c_dv = chart_density_velocity(density_df, f"Densidad — velocidad   N={_size_label(DENSITY_EXPERIMENT_N)}", "fig_density_velocity.png") if not density_df.empty else None

    print("\nBuilding workbook...")
    wb = Workbook()
    if wb.active: wb.remove(wb.active)

    write_sheet(wb, "0. Secuencial",
                "Etapa 0 — Comparación de variantes secuenciales",
                serial_impls, all_reps, avg_data,
                Impl(SEQ, 0), sizes, ct0, cs0)

    write_sheet(wb, "1. OMP base",
                f"Etapa 1 — {best_seq.short_label} vs OpenMP Base",
                [best_seq]+omp_impls, all_reps, avg_data,
                best_seq, sizes, ct1, cs1, csc1)

    write_sheet(wb, "2. OMP compilador",
                f"Etapa 2 — {best_seq.short_label} vs OpenMP + Compiler Opt",
                [best_seq]+omp_opt_impls, all_reps, avg_data,
                best_seq, sizes, ct2, cs2, csc2)

    write_sheet(wb, "3. OMP comp+mem",
                f"Etapa 3 — {best_seq.short_label} vs OpenMP + Compiler & Memory Opt",
                [best_seq]+omp_mem_opt_impls, all_reps, avg_data,
                best_seq, sizes, ct3, cs3, csc3)

    write_sheet(wb, "4. Final",
                "Etapa 4 — Mejor de cada estrategia",
                final_impls, all_reps, avg_data,
                best_seq, sizes, ct4, cs4)

    if not density_df.empty:
        write_density_sheet(wb, density_df, c_dt, c_dv)

    wb.save(OUTPUT_PATH)
    print(f"\nDone.\n  XLSX: {OUTPUT_PATH}\n  Charts: {CHARTS_DIR}/")

if __name__ == "__main__":
    main()
